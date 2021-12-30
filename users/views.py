from django.http import response
from django.http.response import Http404, HttpResponse
from django.shortcuts import render, redirect

from dashboard.views import reverse
from . forms import ChangePasswordForm, SignupForm, LoginForm, LoginHVNForm, ChangePasswordHVNForm
from . models import NewUser, SalePerson, HVN
from django.contrib.auth.models import Group
from django.contrib.auth.decorators import login_required
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.hashers import check_password, make_password, PBKDF2SHA1PasswordHasher
from django.contrib.auth.views import LoginView
import openpyxl
from django.urls import reverse_lazy
from django.contrib.auth import authenticate, login
from django.contrib import messages
from .decorators import unauthenticated_user, unauthenticated_user_HVN
from django.contrib.auth.forms import PasswordChangeForm
from outlet.models import Campain, outletInfo
# Create your views here.

def index(request):
    return HttpResponse('ok')
	



login_required
def PasswordChange(request):
	user = request.user
	if request.method == 'POST':
		form = ChangePasswordForm(request.POST)
		if form.is_valid():
			new_password = form.cleaned_data.get('new_password')
			user.set_password(new_password)
			user.save()
			update_session_auth_hash(request, user)
			return redirect('PasswordChangeDone')
	else:
		form = ChangePasswordForm()

		context = {
			'form':form,
		}

		return render(request, 'users/changepass.html', context)

login_required
def PasswordChangeHVN(request):
	#user = request.user.is_HVN
	user = request.user
	currentpassword = request.user.password
	print(request.user.password)
	a = 'pbkdf2_sha256$260000$tUWlkCUc2UxQ3m5dYOMrt9$/Nf8jFH7XpQ7B3jsW0p2E74Kjs1743fsxggsvPSW6YM='
	check = False
	if currentpassword == a:
		check = True
	
	#Bluesun2021
	if check:	
		if request.method == 'POST':
			form = ChangePasswordHVNForm(request.POST)
			if form.is_valid():
				new_password = form.cleaned_data.get('new_password')
				confirm_password = form.cleaned_data.get('confirm_password')
				if new_password == confirm_password:
					user.set_password(new_password)
					user.save()
					update_session_auth_hash(request, user)
					return redirect("dashboard", campainID = 4)
				return render(request, 'users/changepassHVN.html', {'form':form})
		else:
			form = ChangePasswordHVNForm()
			return render(request, 'users/changepassHVN.html', {'form':form})
	else:
		return redirect("dashboard", campainID = 4)

@unauthenticated_user
def loginPage(request):
    if request.method == 'POST':

        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('page')
        else:
            messages.info(request, "usename or password is incorrect")
            return redirect('loginPage')
    form = LoginForm()
    return render(request, 'users/login.html', {'form':form})

@unauthenticated_user_HVN
def loginHVN(request):
    if request.method == 'POST':

        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect("passwordchangeHVN")
        else:
            messages.info(request, "usename or password is incorrect")
            return redirect('loginHVN')
    form = LoginHVNForm()
    return render(request, 'users/loginHVN.html', {'form':form})

login_required
def page_user(request):
	if request.user.is_salePerson:
		return redirect('listoutlet')
	if request.user.is_HVN:
		return redirect('dashboardforlogin')
	return  Http404
	
	
login_required
def PasswordChangeDone(request):
	return render(request, 'users/successpass.html') 



#import excel data ----> create account for SP 
def upload_user(request):
	if "GET" == request.method:
			return render(request, 'users/upload-user.html', {})
	else:
		excel_file = request.FILES["upload"]
		wb = openpyxl.load_workbook(excel_file)
		
		sheets = wb.sheetnames
		print(sheets[3])
		worksheet = wb[sheets[3]]   #Trang tính

		excel_data = list()
	
		for row in worksheet.iter_rows():
			row_data = list()

			for cell in row:
				#if not (cell.value) == None: 
				row_data.append(str(cell.value))

			excel_data.append(row_data)
		print(len(excel_data)-2)
		
		for i in range(len(excel_data)-1):
			campain = Campain.objects.get(id=6)
			print(excel_data[i+1][2])
			print(excel_data[i+1][3])
			#SPPP
			# outlet = outletInfo.objects.filter(compain = campain)
			# try:
			# 	user1 = NewUser.objects.create_user_SP(user_name=excel_data[i+1][2], password=excel_data[i+1][3])
			# 	sp=SalePerson.objects.create(user=user1, brand = campain, outlet = outlet[1])
			# 	sp.save()
			# except:
			# 	pass
			
			#HVN
			try:
				user1 = NewUser.objects.create_user_HVN(user_name=excel_data[i+1][2] + '@hnk.com', password=excel_data[i+1][3])
				sp=HVN.objects.create(user=user1)
				sp.brand.add(campain)
				sp.save()
			except:	
				pass
		return redirect('PasswordChangeDone')

	
# tigerTP 1
# tigerFA 2
# tigerHZA 3
# heineken 4
# heineken_hnk 5
# STB 6
# bivina 7
# Larue 8
# Larue_SPE 9
