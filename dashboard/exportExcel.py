import datetime
import xlwt
#from outlet.models import tableReport, report_sale
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.chart import (
    PieChart,
    Reference,
    BarChart, Series, Reference,
)
from outlet.models import tableReport, Campain, giftReport, consumerApproachReport, report_sale
from . charts import pie_chart, VOLUME_PERFORMANCE, activation_progress, top10_outlet, gift
from openpyxl.chart.series import DataPoint
from openpyxl.writer.excel import save_virtual_workbook
from .rawData import Table_share, sales_volume, consumers_reached_rawdata, gift_rawdata
from users.models import SalePerson

def export_chart(campainID, all_outlet, from_date, to_date, value, array_image, array_outlet, array_created):  #rawdate = 3, DB =2, full=1
    if value == '3':
        return export_rawdata(campainID, all_outlet, from_date, to_date)
    if value == '4':
        #return export_rawdata(campainID, all_outlet, from_date, to_date)
        return download_files(array_image, array_outlet, array_created, campainID, from_date, to_date)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    #response['Content-Disposition'] = 'attachment;filename={}.xlsx'.format(str(datetime.now()))
    Cp = Campain.objects.get(id = campainID)
    response['Content-Disposition'] = 'attachment;filename={}-{}-{}.xlsx'.format(Cp, from_date, to_date)
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(index = 1 , title = "export-chart")

    ##activation
    Cp = Campain.objects.get(id = campainID)
    table_rp = tableReport.objects.filter(created__gte=from_date, campain = Cp).filter(created__lte=to_date, campain = Cp)
    pie=pie_chart(campainID, table_rp)
    volume_per = VOLUME_PERFORMANCE(campainID, all_outlet)
    activation = activation_progress(campainID, all_outlet)
    top10 = top10_outlet(campainID, all_outlet)
    list_gift_rp = giftReport.objects.filter(created__gte=from_date, campain = Cp).filter(created__lte=to_date, campain = Cp)
    gift_rp = gift(campainID, list_gift_rp)
    ## Volume
    rows = [
        ('Title', 'Average Brand Volume', 'Average Target Volume'),
        ('Average Brand Volume', volume_per[2], 0),
        ('Average Target Volume', volume_per[3], 0),
    ]

    for row1 in rows:
        ws.append(row1)


    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "AVERAGE PERFORMANCE PER ACT"

    data1 = Reference(ws, min_col=2, min_row=1, max_row=3, max_col=3)
    cats = Reference(ws, min_col=1, min_row=2, max_row=3)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats)
    #chart1.shape = 4
    chart1.type = "col"
    #chart1.style = 12
    chart1.grouping = "stacked"
    #chart1.overlap = 100
    ws.add_chart(chart1, "J1")

    ###############################
    data = [
        ['Pie', 'Sold'],
        ['HVN', pie[0]],
        ['Brand', pie[1]],
        ['Other', pie[2]],
        ['Other beer', pie[3]],
    ]

    for row in data:
        ws.append(row)

    pie = PieChart()
    labels = Reference(ws, min_col=1, min_row=5, max_row=8)
    data = Reference(ws, min_col=2, min_row=4, max_row=8)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "TABLE SHARE PERFORMANCE"

    ws.add_chart(pie, "S1")

    ###################################################
    #ACtual
    rows = [
        ('Title', 'Actual Acts', 'Total Acts'),
        ('Total Acts', activation[1], 0),
        ('Actual Acts', activation[0], 0),
    ]

    for row1 in rows:
        ws.append(row1)
    chart3 = BarChart()
    chart3.type = "bar"
    chart3.style = 10
    chart3.title = "ACTIVATION PROGRESS"

    data1 = Reference(ws, min_col=2, min_row=9, max_row=12, max_col=3)
    cats = Reference(ws, min_col=1, min_row=10, max_row=12)
    chart3.add_data(data1, titles_from_data=True)
    chart3.set_categories(cats)
    chart3.type = "bar"
    #chart1.style = 12
    chart3.grouping = "stacked"
    #chart1.overlap = 100
    ws.add_chart(chart3, "A1")
    #gift
    gift5 = 'Gift'
    gift_value5 = 0 
    gift6 = 'Gift'
    gift_value6 = 0 
    gift7 = 'Gift'
    gift_value7 = 0 
    if campainID == 4 or campainID ==1 or campainID ==2 :
        gift5 = gift_rp[1][4]
        gift6 = gift_rp[1][5]
        gift_value5 = gift_rp[0][4]
        gift_value6 = gift_rp[0][5]
    if campainID == 2:
        gift7 = gift_rp[1][6]
        gift_value7 = gift_rp[0][6]
    rows = [
        ('Title', ),
        (gift_rp[1][0], gift_rp[0][0], 0),
        (gift_rp[1][1], gift_rp[0][1], 0),
        (gift_rp[1][2], gift_rp[0][2], 0),
        (gift_rp[1][3], gift_rp[0][3], 0),
        (gift5, gift_value5, 0),
        (gift6, gift_value6, 0),
        (gift7, gift_value7, 0),
    ]

    for row1 in rows:
        ws.append(row1)

    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 10
    #chart2.title = "AVERAGE PERFORMANCE PER ACT"

    data2 = Reference(ws, min_col=2, min_row=12, max_row=20, max_col=3)
    cats = Reference(ws, min_col=1, min_row=13, max_row=19)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats)
    #chart1.shape = 4
    chart2.type = "col"
    chart2.style = 5
    chart2.grouping = "stacked"
    chart2.overlap = 100
    ws.add_chart(chart2, "AB1")
    ######################################################
    #VOLUME PERFORMANCE
    rows = [
        ('Title', 'ActualVolume', 'Target Volume'),
        ('Target Volume', volume_per[1], 0),
        ('ActualVolume', volume_per[0], 0),
    ]

    for row1 in rows:
        ws.append(row1)
    chart4 = BarChart()
    chart4.type = "bar"
    chart4.style = 10
    chart4.title = "VOLUME PERFORMANCE"

    data1 = Reference(ws, min_col=2, min_row=20, max_row=22, max_col=3)
    cats = Reference(ws, min_col=1, min_row=21, max_row=22)
    chart4.add_data(data1, titles_from_data=True)
    chart4.set_categories(cats)
    chart4.type = "bar"
    #chart1.style = 12
    chart4.grouping = "stacked"
    #chart1.overlap = 100
    ws.add_chart(chart4, "A16")
    # TOP 10

    rows = [
       ('Title', '[Brand] Volume', '[Brand] Table Share'),
    ]
    # ('Title', ),
        
    # ('Gift1', 50, 0),
    # ('Gift2', 25, 0),
    # ('Gift3', 35, 0),
    # ('Gift4', 45, 0),
    # ('Gift5', 55, 0),
    # ('Gift6', 65, 0),
    # ('Gift7', 65, 0),
    for i in range(len(top10[2])):
        
        rows.append((top10[2][i],top10[0][i],top10[1][i]))

    for row1 in rows:
        ws.append(row1)

    chart7 = BarChart()
    chart7.type = "col"
    chart7.style = 10
    #chart2.title = "AVERAGE PERFORMANCE PER ACT"

    data2 = Reference(ws, min_col=2, min_row=23, max_row=100, max_col=3)
    cats = Reference(ws, min_col=1, min_row=24, max_row=100)
    chart7.add_data(data2, titles_from_data=True)
    chart7.set_categories(cats)
    
    #chart7.type = "col"
    chart7.style = 5
    #chart7.grouping = "stacked"
    chart7.shape = 4
    #chart7.overlap = 100
    chart7.width = 500
    chart7.height = 25
    if len(top10[2]) < 15:
        chart7.width = 100
        chart7.height = 15
    
    ws.add_chart(chart7, "A31")
    #--------------------------------------------------------> check value
    if value == '2':
        response['Content-Disposition'] = 'attachment;filename={}-DB-{}-{}.xlsx'.format(Cp, from_date, to_date)
        wb.save(response) 
        return response
    #######################------------------------------------------------------------------------------->>>>>>>>>>>>>>>
    # Raw-data
    #wb.save("chart5.xlsx")
    #ws1 = wb.active
    ws1 = wb.create_sheet(index = 0 , title = "raw-data")
    if campainID == 1:
        colums = ['Null', 'Date', 'Province', 'Outlet ID', 'Type', 'Area', 'Address', 'Outlet name', 'Brand Volume Sales', 'HVN Volume Sales', 'Compertion Volume Sales', 'Total Table', 'Brand Table', 'HVN Table', 'Other Beer Table','Other Table', '%Table Share', 'Total Consumers', 'No.Consumers Approached', '% Consumers Reach', 'No. Consumers Bought', '%Conversion', 'Ly 30cl','Ly 33cl 3D','Ly Casablanca', 'Ví', 'Nón Tiger Crystal', 'Voucher Bia', 'Ly 30cl','Ly 33cl 3D','Ly Casablanca', 'Ví', 'Nón Tiger Crystal', 'Voucher Bia']

    elif campainID == 2:
        colums = ['Null', 'Date', 'Province', 'Outlet ID', 'Type', 'Area', 'Address', 'Outlet name', 'Brand Volume Sales', 'HVN Volume Sales', 'Compertion Volume Sales', 'Total Table', 'Brand Table', 'HVN Table', 'Other Beer Table','Other Table', '%Table Share', 'Total Consumers', 'No.Consumers Approached', '% Consumers Reach', 'No. Consumers Bought', '%Conversion','Ly 30cl','Voucher beer', 'Festive Box', 'Túi du lịch Tiger', 'Loa Tiger', 'Ví Tiger ', 'Iphone 13','Ly 30cl','Voucher beer', 'Festive Box', 'Túi du lịch Tiger', 'Loa Tiger', 'Ví Tiger ', 'Iphone 13']

    elif campainID == 4:
        colums = ['Null', 'Date', 'Province', 'Outlet ID', 'Type', 'Area', 'Address', 'Outlet name', 'Brand Volume Sales', 'HVN Volume Sales', 'Compertion Volume Sales', 'Total Table', 'Brand Table', 'HVN Table', 'Other Beer Table','Other Table', '%Table Share', 'Total Consumers', 'No.Consumers Approached', '% Consumers Reach', 'No. Consumers Bought', '%Conversion','Pin sạc','Ba lô', 'Bình Nước', 'Áo thun', 'Loa Bluetooth', 'Ly', 'Pin sạc','Ba lô', 'Bình Nước', 'Áo thun', 'Loa Bluetooth', 'Ly']

    elif campainID == 5:
        colums = ['Null', 'Date', 'Province', 'Outlet ID', 'Type', 'Area', 'Address', 'Outlet name', 'Brand Volume Sales', 'HVN Volume Sales', 'Compertion Volume Sales', 'Total Table', 'Brand Table', 'HVN Table', 'Other Beer Table','Other Table', '%Table Share', 'Total Consumers', 'No.Consumers Approached', '% Consumers Reach', 'No. Consumers Bought', '%Conversion', 'Heineken Alu', 'Ba lô', 'Combo Thời Trang', 'Combo Du Lịch', 'Heineken Alu', 'Ba lô', 'Combo Thời Trang', 'Combo Du Lịch']
    
    elif campainID == 6:
        colums = ['Null', 'Date', 'Province', 'Outlet ID', 'Type', 'Area', 'Address', 'Outlet name', 'Brand Volume Sales', 'HVN Volume Sales', 'Compertion Volume Sales', 'Total Table', 'Brand Table', 'HVN Table', 'Other Beer Table','Other Table', '%Table Share', 'Total Consumers', 'No.Consumers Approached', '% Consumers Reach', 'No. Consumers Bought', '%Conversion', 'Nón Strongbow', 'Túi Jute Bag', 'Túi Canvas ', 'Dù SB', 'Nón Strongbow', 'Túi Jute Bag', 'Túi Canvas ', 'Dù SB']

    elif campainID == 7:
        colums = ['Null','Date', 'Province', 'Outlet ID', 'Type', 'Area', 'Address', 'Outlet name', 'Brand Volume Sales', 'HVN Volume Sales', 'Compertion Volume Sales', 'Total Table', 'Brand Table', 'HVN Table', 'Other Beer Table','Other Table', '%Table Share', 'Total Consumers', 'No.Consumers Approached', '% Consumers Reach', 'No. Consumers Bought', '%Conversion', 'Đồng Hồ Treo Tường', 'Bình Nước 1,6L', 'Ly', 'Túi du lịch', 'Đồng Hồ Treo Tường', 'Bình Nước 1,6L', 'Ly', 'Túi du lịch']
    
    elif campainID == 8:
        colums = ['Null', 'Date', 'Province', 'Outlet ID', 'Type', 'Area', 'Address', 'Outlet name', 'Brand Volume Sales', 'HVN Volume Sales', 'Compertion Volume Sales', 'Total Table', 'Brand Table', 'HVN Table', 'Other Beer Table','Other Table', '%Table Share', 'Total Consumers', 'No.Consumers Approached', '% Consumers Reach', 'No. Consumers Bought', '%Conversion','Áo thun', 'Thùng 12 Lon', 'Nón', 'Ly', 'Áo thun', 'Thùng 12 Lon', 'Nón', 'Ly']
    
    else:
        colums = ['Null', 'Date', 'Province', 'Outlet ID', 'Type', 'Area', 'Address', 'Outlet name', 'Brand Volume Sales', 'HVN Volume Sales', 'Compertion Volume Sales', 'Total Table', 'Brand Table', 'HVN Table', 'Other Beer Table','Other Table', '%Table Share', 'Total Consumers', 'No.Consumers Approached', '% Consumers Reach', 'No. Consumers Bought', '%Conversion','Ba lô','Thùng 12 Lon', 'Nón', '02 Lon Larue', 'Ba lô','Thùng 12 Lon', 'Nón', '02 Lon Larue']
    font_style = xlwt.XFStyle()
    row_num = 1
    for col_num in range(len(colums))[1:]:
        ws1.cell(row_num, col_num, colums[col_num])

    # List date report
    List_date = []
    list_rp = report_sale.objects.filter(created__gte=from_date, campain = Cp).filter(created__lte=to_date, campain = Cp)
    if  not list_rp .exists():
        list_rp  = giftReport.objects.filter(created__gte=from_date, campain = Cp).filter(created__lte=to_date, campain = Cp)
    for date_rp in list_rp:
        if not date_rp.created in List_date:
            List_date.append(date_rp.created)
    #

    sale_person = SalePerson.objects.filter(brand__pk=campainID)  # all_SP
    #if len(List_date) > 0:
    for date_filter in List_date:
        List_outlet = []
        for SP in sale_person:
            outlet=SP.outlet
            rp_table = tableReport.objects.filter(campain = Cp, outlet=outlet, created = date_filter)
            rp_sale =  report_sale.objects.filter(campain=Cp, outlet=outlet, created = date_filter)
            consumers_rp = consumerApproachReport.objects.filter(campain=Cp, outlet=outlet, created = date_filter)
            list_gift_rp = giftReport.objects.filter(campain = Cp, outlet = outlet, created=date_filter)
        
            if rp_table.exists() or rp_sale.exists() or list_gift_rp.exists():
                
                if rp_table.exists() or rp_sale.exists() or list_gift_rp.exists():
                    if not SP.outlet in List_outlet: 
                        List = []
                        total_gift1_receive = 0
                        total_gift2_receive = 0
                        total_gift3_receive = 0
                        total_gift4_receive = 0
                        total_gift5_receive = 0
                        total_gift6_receive = 0
                        total_gift7_receive = 0

                        total_gift1_given = 0
                        total_gift2_given = 0
                        total_gift3_given = 0
                        total_gift4_given = 0
                        total_gift5_given = 0
                        total_gift6_given = 0
                        total_gift7_given = 0
                        List_outlet.append(SP.outlet)
                        table_share = Table_share(campainID, rp_table)
                        SaleVolume = sales_volume(campainID, rp_sale)
                        consumer = consumers_reached_rawdata(campainID, consumers_rp)
                        list_gift_rp = giftReport.objects.filter(campain = Cp, outlet = outlet, created=date_filter)
                        for gift_rp in list_gift_rp:
                            total_gift1_receive = gift_rp.gift1_received
                            total_gift2_receive = gift_rp.gift2_received
                            total_gift3_receive = gift_rp.gift3_received
                            total_gift4_receive = gift_rp.gift4_received
                            total_gift5_receive = gift_rp.gift5_received
                            total_gift6_receive = gift_rp.gift6_received
                            total_gift7_receive = gift_rp.gift7_received

                            total_gift1_given = gift_rp.gift1_given
                            total_gift2_given = gift_rp.gift2_given
                            total_gift3_given = gift_rp.gift3_given
                            total_gift4_given = gift_rp.gift4_given
                            total_gift5_given = gift_rp.gift5_given
                            total_gift6_given = gift_rp.gift6_given
                            total_gift7_given = gift_rp.gift7_given
                        if campainID == 1 or campainID == 4:
                            List = ['Null', date_filter, SP.outlet.province, SP.outlet.ouletID, SP.outlet.type, SP.outlet.area, SP.outlet.outlet_address, SP.outlet.outlet_Name, SaleVolume[0], SaleVolume[1], SaleVolume[2], table_share[0], table_share[1], table_share[2], table_share[3], table_share[4], table_share[5], consumer[0], consumer[1], consumer[2], consumer[3], consumer[4],  total_gift1_receive, total_gift2_receive, total_gift3_receive, total_gift4_receive, total_gift5_receive, total_gift6_receive, total_gift1_given, total_gift2_given, total_gift3_given, total_gift4_given, total_gift5_given, total_gift6_given]
                            row_num +=1
                    
                            for col_num in range(len(List))[1:]:
                                ws1.cell(row_num, col_num, str(List[col_num]))
                        elif campainID == 2:
                            List = ['Null', date_filter, SP.outlet.province, SP.outlet.ouletID, SP.outlet.type, SP.outlet.area, SP.outlet.outlet_address, SP.outlet.outlet_Name, SaleVolume[0], SaleVolume[1], SaleVolume[2], table_share[0], table_share[1], table_share[2], table_share[3], table_share[4], table_share[5], consumer[0], consumer[1], consumer[2], consumer[3], consumer[4], total_gift1_receive, total_gift2_receive, total_gift3_receive, total_gift4_receive, total_gift5_receive, total_gift6_receive,total_gift7_receive, total_gift1_given, total_gift2_given, total_gift3_given, total_gift4_given, total_gift5_given, total_gift6_given, total_gift7_given]
                            row_num +=1
                    
                            for col_num in range(len(List))[1:]:
                                ws1.cell(row_num, col_num, str(List[col_num]))
                        else:
                            #gift_raw_data = gift_rawdata(campainID, list_gift_rp)
                            List = ['Null', date_filter, SP.outlet.province, SP.outlet.ouletID, SP.outlet.type, SP.outlet.area, SP.outlet.outlet_address, SP.outlet.outlet_Name, SaleVolume[0], SaleVolume[1], SaleVolume[2], table_share[0], table_share[1], table_share[2], table_share[3], table_share[4], table_share[5], consumer[0], consumer[1], consumer[2], consumer[3], consumer[4], total_gift1_receive, total_gift2_receive, total_gift3_receive, total_gift4_receive, total_gift1_given, total_gift2_given, total_gift3_given, total_gift4_given]
                            ['Null', 'Date', 'Province', 'Outlet ID', 'Type', 'Area', 'Address', 'Outlet name', 'Brand Volume Sales', 'HVN Volume Sales', 'Compertion Volume Sales', 'Total Table', 'Brand Table', 'HVN Table', 'Other Beer Table','Other Table', '%Table Share', 'Total Consumers', 'No.Consumers Approached', '% Consumers Reach', 'No. Consumers Bought', '%Conversion']
                            row_num +=1
                    
                            for col_num in range(len(List))[1:]:
                                ws1.cell(row_num, col_num, str(List[col_num]))
                    
    #ws1.cell(column=6, row=4).value= 5

    
    wb.save(response) 
    return response

#export_chart()

def export_rawdata(campainID, all_outlet, from_date, to_date):
    Cp = Campain.objects.get(id = campainID)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment;filename={}-RD-{}-{}.xlsx'.format(Cp, from_date, to_date)
    wb = Workbook()
    wb.remove(wb.active)
    ws1 = wb.create_sheet(index = 0 , title = "raw-data")
    
    if campainID == 1:
        colums = ['Null', 'Date', 'Province', 'Outlet ID', 'Type', 'Area', 'Address', 'Outlet name', 'Brand Volume Sales', 'HVN Volume Sales', 'Compertion Volume Sales', 'Total Table', 'Brand Table', 'HVN Table', 'Other Beer Table','Other Table', '%Table Share', 'Total Consumers', 'No.Consumers Approached', '% Consumers Reach', 'No. Consumers Bought', '%Conversion', 'Ly 30cl','Ly 33cl 3D','Ly Casablanca', 'Ví', 'Nón Tiger Crystal', 'Voucher Bia', 'Ly 30cl','Ly 33cl 3D','Ly Casablanca', 'Ví', 'Nón Tiger Crystal', 'Voucher Bia']

    elif campainID == 2:
        colums = ['Null', 'Date', 'Province', 'Outlet ID', 'Type', 'Area', 'Address', 'Outlet name', 'Brand Volume Sales', 'HVN Volume Sales', 'Compertion Volume Sales', 'Total Table', 'Brand Table', 'HVN Table', 'Other Beer Table','Other Table', '%Table Share', 'Total Consumers', 'No.Consumers Approached', '% Consumers Reach', 'No. Consumers Bought', '%Conversion','Ly 30cl','Voucher beer', 'Festive Box', 'Túi du lịch Tiger', 'Loa Tiger', 'Ví Tiger ', 'Iphone 13','Ly 30cl','Voucher beer', 'Festive Box', 'Túi du lịch Tiger', 'Loa Tiger', 'Ví Tiger ', 'Iphone 13']

    elif campainID == 4:
        colums = ['Null', 'Date', 'Province', 'Outlet ID', 'Type', 'Area', 'Address', 'Outlet name', 'Brand Volume Sales', 'HVN Volume Sales', 'Compertion Volume Sales', 'Total Table', 'Brand Table', 'HVN Table', 'Other Beer Table','Other Table', '%Table Share', 'Total Consumers', 'No.Consumers Approached', '% Consumers Reach', 'No. Consumers Bought', '%Conversion','Pin sạc','Ba lô', 'Bình Nước', 'Áo thun', 'Loa Bluetooth', 'Ly', 'Pin sạc','Ba lô', 'Bình Nước', 'Áo thun', 'Loa Bluetooth', 'Ly']

    elif campainID == 5:
        colums = ['Null', 'Date', 'Province', 'Outlet ID', 'Type', 'Area', 'Address', 'Outlet name', 'Brand Volume Sales', 'HVN Volume Sales', 'Compertion Volume Sales', 'Total Table', 'Brand Table', 'HVN Table', 'Other Beer Table','Other Table', '%Table Share', 'Total Consumers', 'No.Consumers Approached', '% Consumers Reach', 'No. Consumers Bought', '%Conversion', 'Heineken Alu', 'Ba lô', 'Combo Thời Trang', 'Combo Du Lịch', 'Heineken Alu', 'Ba lô', 'Combo Thời Trang', 'Combo Du Lịch']
    
    elif campainID == 6:
        colums = ['Null', 'Date', 'Province', 'Outlet ID', 'Type', 'Area', 'Address', 'Outlet name', 'Brand Volume Sales', 'HVN Volume Sales', 'Compertion Volume Sales', 'Total Table', 'Brand Table', 'HVN Table', 'Other Beer Table','Other Table', '%Table Share', 'Total Consumers', 'No.Consumers Approached', '% Consumers Reach', 'No. Consumers Bought', '%Conversion', 'Nón Strongbow', 'Túi Jute Bag', 'Túi Canvas ', 'Dù SB', 'Nón Strongbow', 'Túi Jute Bag', 'Túi Canvas ', 'Dù SB']

    elif campainID == 7:
        colums = ['Null','Date', 'Province', 'Outlet ID', 'Type', 'Area', 'Address', 'Outlet name', 'Brand Volume Sales', 'HVN Volume Sales', 'Compertion Volume Sales', 'Total Table', 'Brand Table', 'HVN Table', 'Other Beer Table','Other Table', '%Table Share', 'Total Consumers', 'No.Consumers Approached', '% Consumers Reach', 'No. Consumers Bought', '%Conversion', 'Đồng Hồ Treo Tường', 'Bình Nước 1,6L', 'Ly', 'Túi du lịch', 'Đồng Hồ Treo Tường', 'Bình Nước 1,6L', 'Ly', 'Túi du lịch']
    
    elif campainID == 8:
        colums = ['Null', 'Date', 'Province', 'Outlet ID', 'Type', 'Area', 'Address', 'Outlet name', 'Brand Volume Sales', 'HVN Volume Sales', 'Compertion Volume Sales', 'Total Table', 'Brand Table', 'HVN Table', 'Other Beer Table','Other Table', '%Table Share', 'Total Consumers', 'No.Consumers Approached', '% Consumers Reach', 'No. Consumers Bought', '%Conversion','Áo thun', 'Thùng 12 Lon', 'Nón', 'Ly', 'Áo thun', 'Thùng 12 Lon', 'Nón', 'Ly']
    
    else:
        colums = ['Null', 'Date', 'Province', 'Outlet ID', 'Type', 'Area', 'Address', 'Outlet name', 'Brand Volume Sales', 'HVN Volume Sales', 'Compertion Volume Sales', 'Total Table', 'Brand Table', 'HVN Table', 'Other Beer Table','Other Table', '%Table Share', 'Total Consumers', 'No.Consumers Approached', '% Consumers Reach', 'No. Consumers Bought', '%Conversion','Ba lô','Thùng 12 Lon', 'Nón', '02 Lon Larue', 'Ba lô','Thùng 12 Lon', 'Nón', '02 Lon Larue']
    font_style = xlwt.XFStyle()
    row_num = 1
    for col_num in range(len(colums))[1:]:
        ws1.cell(row_num, col_num, colums[col_num])

    # List date report
    List_date = []
    list_rp = report_sale.objects.filter(created__gte=from_date, campain = Cp).filter(created__lte=to_date, campain = Cp)
    if  not list_rp .exists():
        list_rp  = giftReport.objects.filter(created__gte=from_date, campain = Cp).filter(created__lte=to_date, campain = Cp)
    for date_rp in list_rp:
        if not date_rp.created in List_date:
            List_date.append(date_rp.created)
    #

    sale_person = SalePerson.objects.filter(brand__pk=campainID)  # all_SP
    #if len(List_date) > 0:
    for date_filter in List_date:
        List_outlet = []
        for SP in sale_person:
            outlet=SP.outlet
            rp_table = tableReport.objects.filter(campain = Cp, outlet=outlet, created = date_filter)
            rp_sale =  report_sale.objects.filter(campain=Cp, outlet=outlet, created = date_filter)
            consumers_rp = consumerApproachReport.objects.filter(campain=Cp, outlet=outlet, created = date_filter)
            list_gift_rp = giftReport.objects.filter(campain = Cp, outlet = outlet, created=date_filter)
        
            if rp_table.exists() or rp_sale.exists() or list_gift_rp.exists():
                
                if rp_table.exists() or rp_sale.exists() or list_gift_rp.exists():
                    if not SP.outlet in List_outlet: 
                        List = []
                        total_gift1_receive = 0
                        total_gift2_receive = 0
                        total_gift3_receive = 0
                        total_gift4_receive = 0
                        total_gift5_receive = 0
                        total_gift6_receive = 0
                        total_gift7_receive = 0

                        total_gift1_given = 0
                        total_gift2_given = 0
                        total_gift3_given = 0
                        total_gift4_given = 0
                        total_gift5_given = 0
                        total_gift6_given = 0
                        total_gift7_given = 0
                        List_outlet.append(SP.outlet)
                        table_share = Table_share(campainID, rp_table)
                        SaleVolume = sales_volume(campainID, rp_sale)
                        consumer = consumers_reached_rawdata(campainID, consumers_rp)
                        list_gift_rp = giftReport.objects.filter(campain = Cp, outlet = outlet, created=date_filter)
                        for gift_rp in list_gift_rp:
                            total_gift1_receive = gift_rp.gift1_received
                            total_gift2_receive = gift_rp.gift2_received
                            total_gift3_receive = gift_rp.gift3_received
                            total_gift4_receive = gift_rp.gift4_received
                            total_gift5_receive = gift_rp.gift5_received
                            total_gift6_receive = gift_rp.gift6_received
                            total_gift7_receive = gift_rp.gift7_received

                            total_gift1_given = gift_rp.gift1_given
                            total_gift2_given = gift_rp.gift2_given
                            total_gift3_given = gift_rp.gift3_given
                            total_gift4_given = gift_rp.gift4_given
                            total_gift5_given = gift_rp.gift5_given
                            total_gift6_given = gift_rp.gift6_given
                            total_gift7_given = gift_rp.gift7_given
                        if campainID == 1 or campainID == 4:
                            List = ['Null', date_filter, SP.outlet.province, SP.outlet.ouletID, SP.outlet.type, SP.outlet.area, SP.outlet.outlet_address, SP.outlet.outlet_Name, SaleVolume[0], SaleVolume[1], SaleVolume[2], table_share[0], table_share[1], table_share[2], table_share[3], table_share[4], table_share[5], consumer[0], consumer[1], consumer[2], consumer[3], consumer[4],  total_gift1_receive, total_gift2_receive, total_gift3_receive, total_gift4_receive, total_gift5_receive, total_gift6_receive, total_gift1_given, total_gift2_given, total_gift3_given, total_gift4_given, total_gift5_given, total_gift6_given]
                            row_num +=1
                    
                            for col_num in range(len(List))[1:]:
                                ws1.cell(row_num, col_num, str(List[col_num]))
                        elif campainID == 2:
                            List = ['Null', date_filter, SP.outlet.province, SP.outlet.ouletID, SP.outlet.type, SP.outlet.area, SP.outlet.outlet_address, SP.outlet.outlet_Name, SaleVolume[0], SaleVolume[1], SaleVolume[2], table_share[0], table_share[1], table_share[2], table_share[3], table_share[4], table_share[5], consumer[0], consumer[1], consumer[2], consumer[3], consumer[4], total_gift1_receive, total_gift2_receive, total_gift3_receive, total_gift4_receive, total_gift5_receive, total_gift6_receive,total_gift7_receive, total_gift1_given, total_gift2_given, total_gift3_given, total_gift4_given, total_gift5_given, total_gift6_given, total_gift7_given]
                            row_num +=1
                    
                            for col_num in range(len(List))[1:]:
                                ws1.cell(row_num, col_num, str(List[col_num]))
                        else:
                            #gift_raw_data = gift_rawdata(campainID, list_gift_rp)
                            List = ['Null', date_filter, SP.outlet.province, SP.outlet.ouletID, SP.outlet.type, SP.outlet.area, SP.outlet.outlet_address, SP.outlet.outlet_Name, SaleVolume[0], SaleVolume[1], SaleVolume[2], table_share[0], table_share[1], table_share[2], table_share[3], table_share[4], table_share[5], consumer[0], consumer[1], consumer[2], consumer[3], consumer[4], total_gift1_receive, total_gift2_receive, total_gift3_receive, total_gift4_receive, total_gift1_given, total_gift2_given, total_gift3_given, total_gift4_given]
                            ['Null', 'Date', 'Province', 'Outlet ID', 'Type', 'Area', 'Address', 'Outlet name', 'Brand Volume Sales', 'HVN Volume Sales', 'Compertion Volume Sales', 'Total Table', 'Brand Table', 'HVN Table', 'Other Beer Table','Other Table', '%Table Share', 'Total Consumers', 'No.Consumers Approached', '% Consumers Reach', 'No. Consumers Bought', '%Conversion']
                            row_num +=1
                    
                            for col_num in range(len(List))[1:]:
                                ws1.cell(row_num, col_num, str(List[col_num]))
                    
    #ws1.cell(column=6, row=4).value= 5

    
    wb.save(response) 
    return response



from zipfile import ZipFile
import os,io,re, magic
from django.http import HttpResponse

# zip folder
def zip_file(array_image, array_outlet, array_created):

    #pathfolder = 'D:/Django_project_API/HNK_project/trackingAPP_project/media/salePerson/'
    pathfolder = 'https://bluesungroup.vn/media/salePerson/'
    #'https://bluesungroup.vn/media/salePerson/'
    abs_src = os.path.abspath(pathfolder)
    outfile = io.BytesIO()
    with ZipFile(outfile,'w') as ivzip:
        for root,subs,files in os.walk(pathfolder):
            for file in files:
                
                filePath = os.path.join(root,file)
                absname = os.path.abspath(filePath)
                arcname = absname[len(abs_src)+1 :] 
                #str_absname = str(absname)
                num_count = 0
                for image in array_image:
                    num_count+=1
                    if str(arcname) in image:
                        print(arcname)
                        arcname = str(array_outlet[num_count]) + '-' + str(array_created[num_count]) + '.png'
                        #arcname = str('ok') + '-' + str('ok') + '.png'
                        print(absname)
                        print(arcname)
                        ivzip.write(absname,arcname)
                
    return outfile

def download_files(array_image, array_outlet, array_created, campainID, from_date, to_date):

    Cp = Campain.objects.get(id = campainID)
    # path  = os.path.join(settings.MEDIA_ROOT+r"\\invoices\\" + contract_id + "\\",year)
    file = zip_file(array_image, array_outlet, array_created)
    response = HttpResponse(file.getvalue(), content_type="application/x-zip-compressed")
    response['Content-Disposition'] = 'attachment;filename={}-PT-{}-{}'.format(Cp,from_date, to_date)+".zip"
    
    return response