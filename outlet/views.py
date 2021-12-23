from django.http.response import Http404
from django.shortcuts import redirect, render, HttpResponse
from django.contrib.auth.decorators import login_required
from .models import overallReport, tableReport, outletInfo, giftReport, search, posmReport, Campain
from .forms import outletInfoForm
from django.views.generic import DetailView, ListView, detail
from users.models import SalePerson
from django.http import JsonResponse
import openpyxl
from django.db.models import Q
# Create your views here.
def index(request):
    return HttpResponse('ok')

login_required
def ListReport(request):
        rp = overallReport.objects.all().count()
        rpview = overallReport.objects.all()
      

        return render(request, 'outlet/index.html',{'tasks': rpview})

login_required
def outlet_create(request, pk):
    if request.method == "POST":
        SP =SalePerson.objects.get(user=request.user)
        form = outletInfoForm(request.POST,is_salePerson=request.user.is_salePerson)
        if form.is_valid():
            outlet_Name = form.cleaned_data.get('outlet_Name')
            type = form.cleaned_data.get('type')
            outlet_address = form.cleaned_data.get('outlet_address')
            ouletID = form.cleaned_data.get('ouletID')
            #create outlet
            p, created = outletInfo.objects.get_or_create(outlet_Name=outlet_Name, type=type,  outlet_address=outlet_address, ouletID=ouletID)
            #CP = Campain.objects.get(program=SP.brand)
            p.compain.add(SP.brand)
            p.save()

            SP.outlet = p
            SP.save()

            return redirect('pagereport')
    else:
        form = outletInfoForm()
        return render(request,"outlet/createinfooutlet.html",{'form':form})
    


# login_required
# class ListOutletView(ListView):
#     model = outletInfo
#     context_object_name = 'list_outlet'
#     #paginate_by = 2
#     template_name = 'outlet/homeoutlet.html'
login_required
def ListOutletView(request):
    user = request.user
    SP = SalePerson.objects.get(user=user)
    compain = SP.brand
    Listoutlet = outletInfo.objects.filter(compain=compain)
    return render(request, 'outlet/homeoutlet.html', {'list_outlet':Listoutlet})

login_required
class OutletDetailView(DetailView):
    model = outletInfo
    template_name = 'outlet/outletdetails.html'
 


login_required
def searchView(request):
    if request.is_ajax():
        outletID = request.POST.get('outletID')
        if outletID != '':
            SP = SalePerson.objects.get(user = request.user)
            all_outlet = outletInfo.objects.filter(compain = SP.brand)
            outletID = request.POST.get('outletID')
            list_outlet = ""
            for outlet in all_outlet:
                if outletID in outlet.ouletID or outletID.lower() in outlet.outlet_Name.lower():
                    list_outlet += f'''<tr>
                                <div class="table-list">
                                        <a class="table-list_name" href={outlet.id}>
                                            {outlet.outlet_Name}
                                        </a>
                                        <p class="table-list_id">
                                            {outlet.ouletID}
                                        </p>
                                        <p class="table-list_address">
                                            {outlet.outlet_address}
                                        </p>
                                </div>
                            </tr>
                            '''
                
            return JsonResponse({'created': 'ok', 'list_outlet':list_outlet})  
        else:
            SP = SalePerson.objects.get(user = request.user)
            district = request.POST.get('district')
            all_outlet = outletInfo.objects.filter(compain = SP.brand)
            list_outlet = ""
            for outlet in all_outlet:
                if district.lower() in outlet.outlet_address.lower():
                    list_outlet += f'''<tr>
                        <div class="table-list">
                                <a class="table-list_name" href={outlet.id}>
                                    {outlet.outlet_Name}
                                </a>
                                <p class="table-list_id">
                                    {outlet.ouletID}
                                </p>
                                <p class="table-list_address">
                                    {outlet.outlet_address}
                                </p>
                        </div>
                    </tr>
                    '''
            #print(list_outlet)
            #search.objects.create(province=province, district=district)
            return JsonResponse({'created': True, 'list_outlet':list_outlet})
    return JsonResponse({'created': False})

login_required
def check_in(request, pk):
    user = request.user
    outlet = outletInfo.objects.get(id=pk)
    SP = SalePerson.objects.get(user=user)
    SP.outlet = outlet
    SP.save()
    return redirect('pagereport')


login_required
def come_back(request, pk):
    pk = str(pk)
    return redirect('https://bluesungroup.vn/outlet/listoutlet/'+pk+'/')

#HVN
def uploadFile_outlet(request):
    try:
        if "GET" == request.method:
            return render(request, 'dashboard/upload-file.html', {})
        else:
            excel_file = request.FILES["upload"]
            wb = openpyxl.load_workbook(excel_file)
            
            sheets = wb.sheetnames
            print(sheets[1])
            worksheet = wb[sheets[1]]   #Trang tính

            excel_data = list()
        
            for row in worksheet.iter_rows():
                row_data = list()

                for cell in row:
                    #if not (cell.value) == None: 
                    row_data.append(str(cell.value))

                excel_data.append(row_data)
            print(len(excel_data)-2)
            # for i in range(len(excel_data)-1):
            #     a = Message(phone_number = '+84'+ excel_data[i+1][0][1:len(excel_data[i+1][0])], content = excel_data[i+1][1])
            #     a.save()
            list_outlet = []
            for i in range(350):
                # campain = Campain.objects.get(id=1)
                # filter_outlet = outletInfo.objects.filter(compain=campain ,ouletID=excel_data[i+1][3], province=excel_data[i+1][2],  outlet_address=excel_data[i+1][6], outlet_Name=excel_data[i+1][7]).count()
                # if filter_outlet <1:
                if excel_data[i+1][2] != None and excel_data[i+1][3] !=None and excel_data[i+1][6] != None:
                    campain = Campain.objects.get(id=2)
                    a = outletInfo.objects.create(created=excel_data[i+1][1], province=excel_data[i+1][2], ouletID=excel_data[i+1][3],
                        type=excel_data[i+1][4], area=excel_data[i+1][5], outlet_address=excel_data[i+1][6], 
                        outlet_Name=excel_data[i+1][7], created_by_HVN = True)
                    
                    a.compain.add(campain)
                    a.save()
                    list_outlet.append(a)
            return render(request, "dashboard/management.html", {'list_outlet':list_outlet})
    except:
        return Http404('file is not support')




# tigerTP 1
# tigerFA 2
# tigerHZA 3
# heineken 4
# heineken_hnk 5
# STB 6
# bivina 7
# Larue 8
# Larue_SPE 9