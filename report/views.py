from os import name
from django.http.response import HttpResponse
from django.shortcuts import render, redirect
from outlet.models import tableReport, report_sale, consumerApproachReport, giftReport, outletInfo, posmReport, overallReport, Campain
from outlet.forms import tableReportForm, reportSaleForm, consumerApproachReportForm, gift_ReceiveReportForm, gift_givenReportForm
from django.views.generic import  DayArchiveView
import datetime
from django.http.response import Http404
import openpyxl
from users.models import SalePerson
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
# Create your views here.




def sum(a, b):
    return str(int(a) + int(b))

def sum_table(a,b,c,d):
    return  str(int(a)+int(b)+int(c)+int(d))

def check_valid_tableReport(a,b,c,d,e):
    f = int(a)+int(b)+int(c)+int(d)
    if f <= int(e):
        return True
    return False

def check_valid_customerReport(a, b):
    if int(a)<=int(b):
        return True
    return False

#---------------------------------------------Report table----------------------------------------------------
login_required
def reportTable(request):

    if request.method == "POST":
        SP = SalePerson.objects.get(user=request.user)
        
        form = tableReportForm(request.POST,is_salePerson=request.user.is_salePerson)
        if form.is_valid():
            other_beer_table = form.cleaned_data.get('other_beer_table')
            brand_table = form.cleaned_data.get('brand_table')
            HVN_table = form.cleaned_data.get('HVN_table')
            other_table = form.cleaned_data.get('other_table')
            # total_table = form.cleaned_data.get('total_table')

            
            report_table = tableReport.objects.filter(created = datetime.date.today(), SP = request.user, outlet = SP.outlet).count()

            if report_table < 1:
                total_table = sum_table(other_beer_table, brand_table, HVN_table, other_table)
                p= tableReport.objects.create(SP=request.user, outlet=SP.outlet, campain=SP.brand, total_table=total_table,
                        other_beer_table=other_beer_table, other_table=other_table, brand_table=brand_table,  HVN_table=HVN_table)
                p.save()
                
                return render(request,"report/create_reportTable.html", {'other_beer_table':other_beer_table, 'brand_table':brand_table,
                                    'HVN_table':HVN_table,'total_table':total_table, 'other_table':other_table})

           
            report_table = tableReport.objects.get(created = datetime.date.today(), SP = request.user, outlet = SP.outlet)
            
            report_table.other_beer_table = other_beer_table
            report_table.other_table = other_table
            report_table.brand_table = brand_table
            report_table.HVN_table = HVN_table 
            total_table = sum_table(other_beer_table, brand_table, HVN_table, other_table)
            report_table.total_table = total_table
            report_table.save()
            
            return render(request,"report/sum-totaltable.html",{'sum_other_beer_table':report_table.other_beer_table, 'other_beer_table':other_beer_table,
                'sum_brand_table':report_table.brand_table, 'brand_table':brand_table, 'sum_HVN_table':report_table.HVN_table, 'HVN_table':HVN_table,
                'sum_total_table': report_table.total_table, 'total_table':total_table, 'sum_other_table':report_table.other_table, 'other_table':other_table})
        
    else:
        form = tableReportForm()
        return render(request,"report/table-number.html",{'form':form})

#--------------------------------------------------Sale report-----------------------------------------------------
login_required
def reportSale(request):
    if request.method == "POST":
        SP = SalePerson.objects.get(user=request.user)
        form = reportSaleForm(request.POST,is_salePerson=request.user.is_salePerson)
        if form.is_valid():
            beer_brand = form.cleaned_data.get('beer_brand')
            beer_HVN = form.cleaned_data.get('beer_HVN')
            beer_other = form.cleaned_data.get('beer_other')
            
            report = report_sale.objects.filter(created = datetime.date.today(), SP = request.user, outlet = SP.outlet).count()

            if report < 1:
                p = report_sale.objects.create(SP=request.user, outlet=SP.outlet, beer_brand=beer_brand, beer_HVN=beer_HVN, beer_other=beer_other, campain=SP.brand)
                p.save()
                return render(request,"report/create_salereport.html", {'beer_brand':beer_brand, 
                                        'beer_HVN':beer_HVN,'beer_other':beer_other})

           
            report = report_sale.objects.get(created = datetime.date.today(), SP = request.user, outlet = SP.outlet)
            
            report.beer_brand = sum(beer_brand, report.beer_brand)
            report.beer_HVN = sum(beer_HVN, report.beer_HVN)
            report.beer_other = sum(beer_other, report.beer_other)
            
            report.save()
            return render(request,"report/report-total-sale.html",{'sum_beer_brand':report.beer_brand, 'beer_brand':beer_brand,
                'sum_beer_HVN': report.beer_HVN, 'beer_HVN':beer_HVN, 'sum_beer_other':report.beer_other, 'beer_other': beer_other})
        
    else:
        form = reportSaleForm()
        SP = SalePerson.objects.get(user=request.user)
        beer_brand = '0'
        beer_other = '0'
        beer_HVN = '0'
        report = report_sale.objects.filter(created = datetime.date.today(), SP = request.user, outlet = SP.outlet).count()
        if report == 1:
            report = report_sale.objects.get(created = datetime.date.today(), SP = request.user, outlet = SP.outlet)
            beer_brand = report.beer_brand
            beer_other = report.beer_other
            beer_HVN = report.beer_HVN
        return render(request,"report/sales.html", {'form':form, 'beer_brand':beer_brand, 'beer_other':beer_other, 'beer_HVN':beer_HVN})

#------------------------------------------------------Report customer----------------------------------------
login_required
def report_customer(request):
    if request.method == "POST":
        form = consumerApproachReportForm(request.POST,is_salePerson=request.user.is_salePerson)
        if form.is_valid():
            SP = SalePerson.objects.get(user=request.user)
            consumers_approach = form.cleaned_data.get('consumers_approach')
            consumers_brough = form.cleaned_data.get('consumers_brough')
            Total_Consumers = form.cleaned_data.get('Total_Consumers')
            if (check_valid_customerReport(consumers_approach, Total_Consumers) == False) or (check_valid_customerReport(consumers_brough, consumers_approach) == False):
                return render(request,"report/alert-customer.html", {'consumers_approach':consumers_approach, 
                            'consumers_brough':consumers_brough,'Total_Consumers':Total_Consumers})

            
            report = consumerApproachReport.objects.filter(created = datetime.date.today(), SP = request.user, outlet = SP.outlet).count()

            if report < 1:
                p= consumerApproachReport.objects.create(SP=request.user, outlet=SP.outlet, campain=SP.brand,
                            consumers_approach=consumers_approach, consumers_brough=consumers_brough, Total_Consumers=Total_Consumers)
                p.save()
                return render(request,"report/create-report-customer.html", {'consumers_approach':consumers_approach, 
                            'consumers_brough':consumers_brough,'Total_Consumers':Total_Consumers})

           
            report = consumerApproachReport.objects.get(created = datetime.date.today(), SP = request.user, outlet = SP.outlet)
            
            report.consumers_approach = sum(consumers_approach, report.consumers_approach)
            report.consumers_brough = sum(consumers_brough, report.consumers_brough)
            report.Total_Consumers = sum(Total_Consumers, report.Total_Consumers)
            
            report.save()
            return render(request,"report/report-total-consumer.html",{'sum_consumers_approach':report.consumers_approach, 
                'consumers_approach':consumers_approach, 'sum_consumers_brough': report.consumers_brough, 'consumers_brough':consumers_brough, 
                'sum_Total_Consumers':report.Total_Consumers, 'Total_Consumers': Total_Consumers})
        
    else:
        form = consumerApproachReportForm()
        SP = SalePerson.objects.get(user=request.user)
        Total_Consumers = '0'
        consumers_approach = '0'
        consumers_brough = '0'
        report = consumerApproachReport.objects.filter(created = datetime.date.today(), SP = request.user, outlet = SP.outlet).count()
        if report == 1:
            report = consumerApproachReport.objects.get(created = datetime.date.today(), SP = request.user, outlet = SP.outlet)
            Total_Consumers = report.Total_Consumers
            consumers_approach = report.consumers_approach
            consumers_brough = report.consumers_brough
        return render(request,"report/customer-access.html", {'form':form, 'Total_Consumers':Total_Consumers, 'consumers_approach':consumers_approach,'consumers_brough':consumers_brough})

#---------------------------------------------Gift report-----------------------------------------------------
login_required
def gift_receiveReport(request):
    if request.method == "POST":
        form = gift_ReceiveReportForm(request.POST,is_salePerson=request.user.is_salePerson)
        if form.is_valid():
            SP = SalePerson.objects.get(user=request.user)
            gift1_received = form.cleaned_data.get('gift1_received')
            gift2_received = form.cleaned_data.get('gift2_received')
            gift3_received = form.cleaned_data.get('gift3_received')   
            gift4_received = form.cleaned_data.get('gift4_received')
            gift5_received = form.cleaned_data.get('gift5_received')
            gift6_received = form.cleaned_data.get('gift6_received')
            gift7_received = form.cleaned_data.get('gift7_received')

            campain7 = Campain.objects.get(id=7)
            campain1 = Campain.objects.get(id=1)
            campain2 = Campain.objects.get(id=2)
            campain3 = Campain.objects.get(id=3)
            campain4 = Campain.objects.get(id=4)
            campain5 = Campain.objects.get(id=5)
            campain6 = Campain.objects.get(id=6)
            campain8 = Campain.objects.get(id=8)
            campain9 = Campain.objects.get(id=9)
            #form = gift_ReceiveReportForm()
            # report for 4 gift
            if SP.brand == campain7 or SP.brand == campain8 or SP.brand == campain5 or SP.brand == campain9 or SP.brand == campain6:
                report = giftReport.objects.filter(created = datetime.date.today(), SP = request.user, outlet = SP.outlet).count()
                if report < 1:
                    p = giftReport.objects.create(SP=request.user, outlet=SP.outlet, campain=SP.brand, gift1_received=gift1_received, 
                            gift2_received=gift2_received, gift3_received=gift3_received, gift4_received=gift4_received)
                    p.save()
                    if SP.brand == campain7:
                        return render(request, "report/create-list-gift-receive.html", {'gift1_received':gift1_received, 
                        'gift2_received':gift2_received,  'gift3_received':gift3_received ,'gift4_received':gift4_received, 'gift_name_1':'Túi du lịch', 'gift_name_2':'Đồng Hồ Treo Tường', 'gift_name_3':'Bình Nước 1,6L', 'gift_name_4':'Ly'})
                    if SP.brand == campain8:
                        return render(request, "report/create-list-gift-receive.html", {'gift1_received':gift1_received, 
                        'gift2_received':gift2_received,  'gift3_received':gift3_received ,'gift4_received':gift4_received, 'gift_name_1':'Áo thun', 'gift_name_2':'Thùng 12 Lon', 'gift_name_3':'Nón', 'gift_name_4':'Ly'})
                    if SP.brand == campain5:
                        return render(request, "report/create-list-gift-receive.html", {'gift1_received':gift1_received, 
                        'gift2_received':gift2_received,  'gift3_received':gift3_received ,'gift4_received':gift4_received, 'gift_name_1':'Heineken Alu', 'gift_name_2':'Ba lô', 'gift_name_3':'Combo Thời Trang', 'gift_name_4':'Combo Du Lịch'})
                    if SP.brand == campain6:
                        return render(request, "report/create-list-gift-receive.html", {'gift1_received':gift1_received, 
                        'gift2_received':gift2_received,  'gift3_received':gift3_received ,'gift4_received':gift4_received, 'gift_name_1':'Nón Strongbow', 'gift_name_2':'Túi Jute Bag', 'gift_name_3':'Túi Canvas ', 'gift_name_4':'Dù SB'})

                    if SP.brand == campain9:
                        return render(request, "report/create-list-gift-receive.html", {'gift1_received':gift1_received, 
                        'gift2_received':gift2_received,  'gift3_received':gift3_received ,'gift4_received':gift4_received, 'gift_name_1':'Ba lô', 'gift_name_2':'Thùng 12 Lon', 'gift_name_3':'Nón', 'gift_name_4':'02 Lon Larue'})
                    
                report = giftReport.objects.get(created = datetime.date.today(), SP = request.user, outlet = SP.outlet)
                report.gift1_received = sum(gift1_received, report.gift1_received)
                report.gift2_received = sum(gift2_received, report.gift2_received)
                report.gift3_received = sum(gift3_received, report.gift3_received)
                report.gift4_received = sum(gift4_received, report.gift4_received)
                report.save()
                if SP.brand == campain7:
                    return render(request, "report/create-list-gift-receive.html", {'gift1_received':report.gift1_received, 'gift2_received':report.gift2_received,'gift3_received':report.gift3_received, 'gift4_received':report.gift4_received, 'gift_name_1':'Túi du lịch', 'gift_name_2':'Đồng Hồ Treo Tường', 'gift_name_3':'Bình Nước 1,6L', 'gift_name_4':'Ly'})
                if SP.brand == campain8:
                        return render(request, "report/create-list-gift-receive.html", {'gift1_received':report.gift1_received, 
                        'gift2_received':report.gift2_received,  'gift3_received':report.gift3_received ,'gift4_received':report.gift4_received, 'gift_name_1':'Áo thun', 'gift_name_2':'Thùng 12 Lon', 'gift_name_3':'Nón', 'gift_name_4':'Ly'})
                if SP.brand == campain5:
                    return render(request, "report/create-list-gift-receive.html", {'gift1_received':report.gift1_received, 
                    'gift2_received':report.gift2_received,  'gift3_received':report.gift3_received ,'gift4_received':report.gift4_received, 'gift_name_1':'Heineken Alu', 'gift_name_2':'Ba lô', 'gift_name_3':'Combo Thời Trang', 'gift_name_4':'Combo Du Lịch'})
                
                if SP.brand == campain6:
                        return render(request, "report/create-list-gift-receive.html", {'gift1_received':report.gift1_received, 
                        'gift2_received':report.gift2_received,  'gift3_received':report.gift3_received ,'gift4_received':report.gift4_received, 'gift_name_1':'Nón Strongbow', 'gift_name_2':'Túi Jute Bag', 'gift_name_3':'Túi Canvas ', 'gift_name_4':'Dù SB'})
                        
                if SP.brand == campain9:
                        return render(request, "report/create-list-gift-receive.html", {'gift1_received':report.gift1_received, 
                        'gift2_received':report.gift2_received,  'gift3_received':report.gift3_received ,'gift4_received':report.gift4_received, 'gift_name_1':'Ba lô', 'gift_name_2':'Thùng 12 Lon', 'gift_name_3':'Nón', 'gift_name_4':'02 Lon Larue'})
            
            if SP.brand == campain4 or SP.brand == campain1:      ###
                report = giftReport.objects.filter(created = datetime.date.today(), SP = request.user, outlet = SP.outlet).count()
                if report < 1:
                    p = giftReport.objects.create(SP=request.user, outlet=SP.outlet, campain=SP.brand, gift1_received=gift1_received, 
                            gift2_received=gift2_received, gift3_received=gift3_received, gift4_received=gift4_received, 
                            gift5_received=gift5_received, gift6_received=gift6_received)
                    p.save()
                    if SP.brand == campain4:
                        return render(request, "list_gift/create-list-gift-receive.html", {'gift1_received':gift1_received, 
                        'gift2_received':gift2_received,  'gift3_received':gift3_received ,'gift4_received':gift4_received, 'gift5_received':gift5_received, 'gift6_received':gift6_received, 'gift_name_1':'Pin sạc', 'gift_name_2':'Ba lô', 'gift_name_3':'Bình Nước', 'gift_name_4':'Áo thun', 'gift_name_5':'Loa Bluetooth', 'gift_name_6':'Ly'})
                    if SP.brand == campain1:
                        return render(request, "list_gift/create-list-gift-receive.html", {'gift1_received':gift1_received, 
                        'gift2_received':gift2_received,  'gift3_received':gift3_received ,'gift4_received':gift4_received, 'gift5_received':gift5_received, 'gift6_received':gift6_received, 'gift_name_1':'Ly 30cl', 'gift_name_2':'Ly 33cl 3D', 'gift_name_3':'Ly Casablanca', 'gift_name_4':'Ví', 'gift_name_5':'Nón Tiger Crystal', 'gift_name_6':'Voucher Bia'})
                report = giftReport.objects.get(created = datetime.date.today(), SP = request.user, outlet = SP.outlet)
                report.gift1_received = sum(gift1_received, report.gift1_received)
                report.gift2_received = sum(gift2_received, report.gift2_received)
                report.gift3_received = sum(gift3_received, report.gift3_received)
                report.gift4_received = sum(gift4_received, report.gift4_received)
                report.gift5_received = sum(gift5_received, report.gift5_received)
                report.gift6_received = sum(gift6_received, report.gift6_received)
                report.save()
                if SP.brand == campain1:
                        return render(request, "list_gift/create-list-gift-receive.html", {'gift1_received':report.gift1_received, 
                        'gift2_received':report.gift2_received,  'gift3_received':report.gift3_received ,'gift4_received':report.gift4_received, 'gift5_received':report.gift5_received, 'gift6_received':report.gift6_received, 'gift_name_1':'Ly 30cl', 'gift_name_2':'Ly 33cl 3D', 'gift_name_3':'Ly Casablanca', 'gift_name_4':'Ví', 'gift_name_5':'Nón Tiger Crystal', 'gift_name_6':'Voucher Bia'})
                if SP.brand == campain4:
                    return render(request, "list_gift/create-list-gift-receive.html", {'gift1_received':report.gift1_received, 'gift2_received':report.gift2_received,'gift3_received':report.gift3_received, 'gift4_received':report.gift4_received, 'gift5_received':report.gift5_received, 'gift6_received':report.gift6_received, 'gift_name_1':'Pin sạc', 'gift_name_2':'Ba lô', 'gift_name_3':'Bình Nước', 'gift_name_4':'Áo thun', 'gift_name_5':'Loa Bluetooth', 'gift_name_6':'Ly' })
            if  SP.brand == campain2:
                report = giftReport.objects.filter(created = datetime.date.today(), SP = request.user, outlet = SP.outlet).count()
                if report < 1:
                    p = giftReport.objects.create(SP=request.user, outlet=SP.outlet, campain=SP.brand, gift1_received=gift1_received, 
                            gift2_received=gift2_received, gift3_received=gift3_received, gift4_received=gift4_received, 
                            gift5_received=gift5_received, gift6_received=gift6_received, gift7_received=gift7_received)
                    p.save()
                    
                    return render(request, "list_gift1/create-list-gift-receive.html", {'gift1_received':gift1_received, 
                        'gift2_received':gift2_received,  'gift3_received':gift3_received ,'gift4_received':gift4_received, 'gift5_received':gift5_received, 'gift6_received':gift6_received, 'gift7_received':gift7_received,'gift_name_1':'Ly 30cl', 'gift_name_2':'Voucher beer', 'gift_name_3':'Festive Box', 'gift_name_4':'Túi di lịch Tiger', 'gift_name_5':'Loa Tiger', 'gift_name_6':'Ví Tiger ', 'gift_name_7':'Iphone 13'})
            
                report = giftReport.objects.get(created = datetime.date.today(), SP = request.user, outlet = SP.outlet)
                report.gift1_received = sum(gift1_received, report.gift1_received)
                report.gift2_received = sum(gift2_received, report.gift2_received)
                report.gift3_received = sum(gift3_received, report.gift3_received)
                report.gift4_received = sum(gift4_received, report.gift4_received)
                report.gift5_received = sum(gift5_received, report.gift5_received)
                report.gift6_received = sum(gift6_received, report.gift6_received)
                report.gift7_received = sum(gift6_received, report.gift7_received)
                report.save()
                return render(request, "list_gift1/create-list-gift-receive.html", {'gift1_received':report.gift1_received, 
                        'gift2_received':report.gift2_received,  'gift3_received':report.gift3_received ,'gift4_received':report.gift4_received, 'gift5_received':report.gift5_received, 'gift6_received':report.gift6_received,'gift7_received':report.gift7_received, 'gift_name_1':'Ly 30cl', 'gift_name_2':'Voucher beer', 'gift_name_3':'Festive Box', 'gift_name_4':'Túi di lịch Tiger', 'gift_name_5':'Loa Tiger', 'gift_name_6':'Ví Tiger ', 'gift_name_7':'Iphone 13'})
            if  SP.brand == campain3:
                report = giftReport.objects.filter(created = datetime.date.today(), SP = request.user, outlet = SP.outlet).count()
                if report < 1:
                    p = giftReport.objects.create(SP=request.user, outlet=SP.outlet, campain=SP.brand, gift1_received=gift1_received, 
                            gift2_received=gift2_received, gift3_received=gift3_received)
                    p.save()
                    
                    return render(request, "list_gift2/create-list-gift-receive.html", {'gift1_received':gift1_received, 
                        'gift2_received':gift2_received,  'gift3_received':gift3_received ,'gift_name_1':'E-voucher 25k', 'gift_name_2':'E-voucher 50k', 'gift_name_3':'E-voucher 100k'})
                report = giftReport.objects.get(created = datetime.date.today(), SP = request.user, outlet = SP.outlet)
                report.gift1_received = sum(gift1_received, report.gift1_received)
                report.gift2_received = sum(gift2_received, report.gift2_received)
                report.gift3_received = sum(gift3_received, report.gift3_received)
                report.save()
                return render(request, "list_gift2/create-list-gift-receive.html", {'gift1_received':report.gift1_received, 
                        'gift2_received':report.gift2_received,  'gift3_received':report.gift3_received,'gift_name_1':'E-voucher 25k', 'gift_name_2':'E-voucher 50k', 'gift_name_3':'E-voucher 100k'})
    else:
        SP = SalePerson.objects.get(user=request.user)
        campain7 = Campain.objects.get(id=7)
        campain4 = Campain.objects.get(id=4)
        campain8 = Campain.objects.get(id=8)
        campain5 = Campain.objects.get(id=5)
        campain1 = Campain.objects.get(id=1)
        campain2 = Campain.objects.get(id=2)
        campain3 = Campain.objects.get(id=3)
        campain6 = Campain.objects.get(id=6)
        campain9 = Campain.objects.get(id=9)

        form = gift_ReceiveReportForm()
        if SP.brand == campain7:
            return render(request,"report/listgift-received.html", {'form':form, 'gift_name_1':'Túi du lịch', 'gift_name_2':'Đồng Hồ Treo Tường', 'gift_name_3':'Bình Nước 1,6L', 'gift_name_4':'Ly'})
        elif SP.brand == campain8:
            return render(request,"report/listgift-received.html", {'form':form, 'gift_name_1':'Áo thun', 'gift_name_2':'Thùng 12 Lon', 'gift_name_3':'Nón', 'gift_name_4':'Ly'})

        elif SP.brand == campain9:
            return render(request,"report/listgift-received.html", {'form':form, 'gift_name_1':'Ba lô', 'gift_name_2':'Thùng 12 Lon', 'gift_name_3':'Nón', 'gift_name_4':'02 Lon Larue'})

        elif SP.brand == campain1:
            return render(request,"list_gift/listgift-received.html", {'form':form, 'gift_name_1':'Ly 30cl', 'gift_name_2':'Ly 33cl 3D', 'gift_name_3':'Ly Casablanca', 'gift_name_4':'Ví', 'gift_name_5':'Nón Tiger Crystal', 'gift_name_6':'Voucher Bia'})

        elif SP.brand == campain2:
            return render(request,"list_gift1/listgift-received.html", {'form':form, 'gift_name_1':'Ly 30cl', 'gift_name_2':'Voucher beer', 'gift_name_3':'Festive Box', 'gift_name_4':'Túi di lịch Tiger', 'gift_name_5':'Loa Tiger', 'gift_name_6':'Ví Tiger ', 'gift_name_7':'Iphone 13'})

        elif SP.brand == campain3:
            return render(request,"list_gift2/listgift-received.html", {'form':form, 'gift_name_1':'E-voucher 25k', 'gift_name_2':'E-voucher 50k', 'gift_name_3':'E-voucher 100k'})

        elif SP.brand == campain4:
            return render(request,"list_gift/listgift-received.html", {'form':form, 'gift_name_1':'Pin sạc', 'gift_name_2':'Ba lô', 'gift_name_3':'Bình Nước', 'gift_name_4':'Áo thun', 'gift_name_5':'Loa Bluetooth', 'gift_name_6':'Ly'})
        elif SP.brand == campain5:
            return render(request,"report/listgift-received.html", {'form':form, 'gift_name_1':'Heineken Alu', 'gift_name_2':'Ba lô', 'gift_name_3':'Combo Thời Trang', 'gift_name_4':'Combo Du Lịch'})
        elif SP.brand == campain6:
            return render(request,"report/listgift-received.html", {'form':form, 'gift_name_1':'Nón Strongbow', 'gift_name_2':'Túi Jute Bag', 'gift_name_3':'Túi Canvas ', 'gift_name_4':'Dù SB'})
        else:
            return render(request,"report/listgift-received.html", {'form':form})
        

login_required
def gift_givenReport(request):
    try:
        if request.method == "POST":
            form = gift_givenReportForm(request.POST,is_salePerson=request.user.is_salePerson)
            if form.is_valid():
                gift1_given = form.cleaned_data.get('gift1_given')
                gift2_given = form.cleaned_data.get('gift2_given')
                gift3_given = form.cleaned_data.get('gift3_given') 
                gift4_given = form.cleaned_data.get('gift4_given') 
                gift5_given = form.cleaned_data.get('gift5_given') 
                gift6_given = form.cleaned_data.get('gift6_given')
                gift7_given = form.cleaned_data.get('gift7_given')
                campain7 = Campain.objects.get(id=7)
                campain4 = Campain.objects.get(id=4)
                campain8 = Campain.objects.get(id=8)
                campain5 = Campain.objects.get(id=5)
                campain1 = Campain.objects.get(id=1)
                campain2 = Campain.objects.get(id=2)
                campain3 = Campain.objects.get(id=3)
                campain6 = Campain.objects.get(id=6)
                campain9 = Campain.objects.get(id=9)
                form = gift_ReceiveReportForm()
                SP = SalePerson.objects.get(user=request.user)
                if SP.brand == campain6 or SP.brand == campain7 or SP.brand == campain8 or SP.brand == campain5 or SP.brand == campain9:
                    SP = SalePerson.objects.get(user=request.user)
                    report = giftReport.objects.get(created = datetime.date.today(), SP = request.user, outlet = SP.outlet)
                    report.gift1_given = sum(gift1_given, report.gift1_given)
                    report.gift2_given = sum(gift2_given, report.gift2_given)
                    report.gift3_given = sum(gift3_given, report.gift3_given)
                    report.gift4_given = sum(gift4_given, report.gift4_given)
                    report.save()
                    return redirect('quantity-gift')
                if SP.brand == campain4 or SP.brand == campain1:
                    SP = SalePerson.objects.get(user=request.user)
                    report = giftReport.objects.get(created = datetime.date.today(), SP = request.user, outlet = SP.outlet)
                    report.gift1_given = sum(gift1_given, report.gift1_given)
                    report.gift2_given = sum(gift2_given, report.gift2_given)
                    report.gift3_given = sum(gift3_given, report.gift3_given)
                    report.gift4_given = sum(gift4_given, report.gift4_given)
                    report.gift5_given = sum(gift5_given, report.gift5_given)
                    report.gift6_given = sum(gift6_given, report.gift6_given)
                    report.save()
                    return redirect('quantity-gift')
                if SP.brand == campain2:
                    SP = SalePerson.objects.get(user=request.user)
                    report = giftReport.objects.get(created = datetime.date.today(), SP = request.user, outlet = SP.outlet)
                    report.gift1_given = sum(gift1_given, report.gift1_given)
                    report.gift2_given = sum(gift2_given, report.gift2_given)
                    report.gift3_given = sum(gift3_given, report.gift3_given)
                    report.gift4_given = sum(gift4_given, report.gift4_given)
                    report.gift5_given = sum(gift5_given, report.gift5_given)
                    report.gift6_given = sum(gift6_given, report.gift6_given)
                    report.gift7_given = sum(gift6_given, report.gift7_given)
                    report.save()
                    return redirect('quantity-gift')
                if SP.brand == campain3:
                    SP = SalePerson.objects.get(user=request.user)
                    report = giftReport.objects.get(created = datetime.date.today(), SP = request.user, outlet = SP.outlet)
                    report.gift1_given = sum(gift1_given, report.gift1_given)
                    report.gift2_given = sum(gift2_given, report.gift2_given)
                    report.gift3_given = sum(gift3_given, report.gift3_given)
                    report.save()
                    return redirect('quantity-gift')
        else:
            form = gift_givenReportForm()
            campain7 = Campain.objects.get(id=7)
            campain4 = Campain.objects.get(id=4)
            campain8 = Campain.objects.get(id=8)
            campain5 = Campain.objects.get(id=5)
            campain1 = Campain.objects.get(id=1)
            campain2 = Campain.objects.get(id=2)
            campain3 = Campain.objects.get(id=3)
            campain6 = Campain.objects.get(id=6)
            campain9 = Campain.objects.get(id=9)
            #form = gift_ReceiveReportForm()
            SP = SalePerson.objects.get(user=request.user)
            if SP.brand == campain7:
                return render(request,"report/listgift-sent.html", {'form':form, 'gift_name_1':'Túi du lịch', 'gift_name_2':'Đồng Hồ Treo Tường', 'gift_name_3':'Bình Nước 1,6L', 'gift_name_4':'Ly'})
            if SP.brand == campain8:
                return render(request,"report/listgift-sent.html", {'form':form, 'gift_name_1':'Áo thun', 'gift_name_2':'Thùng 12 Lon', 'gift_name_3':'Nón', 'gift_name_4':'Ly'})
            
            if SP.brand == campain9:
                return render(request,"report/listgift-sent.html", {'form':form, 'gift_name_1':'Ba lô', 'gift_name_2':'Thùng 12 Lon', 'gift_name_3':'Nón', 'gift_name_4':'02 Lon Larue'})
            
            if SP.brand == campain5:
                return render(request,"report/listgift-sent.html", {'form':form, 'gift_name_1':'Heineken Alu', 'gift_name_2':'Ba lô', 'gift_name_3':'Combo Thời Trang', 'gift_name_4':'Combo Du Lịch'})
            
            if SP.brand == campain6:
                return render(request,"report/listgift-sent.html", {'form':form, 'gift_name_1':'Nón Strongbow', 'gift_name_2':'Túi Jute Bag', 'gift_name_3':'Túi Canvas ', 'gift_name_4':'Dù SB'})

            if SP.brand == campain1:
                return render(request,"list_gift/listgift-sent.html", {'form':form, 'gift_name_1':'Ly 30cl', 'gift_name_2':'Ly 33cl 3D', 'gift_name_3':'Ly Casablanca', 'gift_name_4':'Ví', 'gift_name_5':'Nón Tiger Crystal', 'gift_name_6':'Voucher Bia'})

            if SP.brand == campain2:
                return render(request,"list_gift1/listgift-sent.html", {'form':form, 'gift_name_1':'Ly 30cl', 'gift_name_2':'Voucher beer', 'gift_name_3':'Festive Box', 'gift_name_4':'Túi di lịch Tiger', 'gift_name_5':'Loa Tiger', 'gift_name_6':'Ví Tiger ', 'gift_name_7':'Iphone 13'})

            if SP.brand == campain3:
                return render(request,"list_gift2/listgift-sent.html", {'form':form, 'gift_name_1':'E-voucher 25k', 'gift_name_2':'E-voucher 50k', 'gift_name_3':'E-voucher 100k'})

            if SP.brand == campain4:
                return render(request,"list_gift/listgift-sent.html", {'form':form, 'gift_name_1':'Pin sạc', 'gift_name_2':'Ba lô', 'gift_name_3':'Bình Nước', 'gift_name_4':'Áo thun', 'gift_name_5':'Loa Bluetooth', 'gift_name_6':'Ly'})
    except:
        
        campain7 = Campain.objects.get(id=7)
        campain4 = Campain.objects.get(id=4)
        campain8 = Campain.objects.get(id=8)
        campain5 = Campain.objects.get(id=5)
        campain1 = Campain.objects.get(id=1)
        campain2 = Campain.objects.get(id=2)
        campain3 = Campain.objects.get(id=3)
        campain6 = Campain.objects.get(id=6)
        campain9 = Campain.objects.get(id=9)
        SP = SalePerson.objects.get(user=request.user)
        if SP.brand == campain7:
            return render(request, 'report/alert-gift-given.html', {'gift1_given':gift1_given, 'gift2_given':gift2_given,
                'gift3_given':gift3_given,'gift4_given':gift4_given, 'gift_name_1':'Túi du lịch', 'gift_name_2':'Đồng Hồ Treo Tường', 'gift_name_3':'Bình Nước 1,6L', 'gift_name_4':'Ly'} )
        if SP.brand == campain8:
            return render(request, 'report/alert-gift-given.html', {'gift1_given':gift1_given, 'gift2_given':gift2_given,
                'gift3_given':gift3_given,'gift4_given':gift4_given, 'gift_name_1':'Áo thun', 'gift_name_2':'Thùng 12 Lon', 'gift_name_3':'Nón', 'gift_name_4':'Ly'})
        
        if SP.brand == campain9:
            return render(request, 'report/alert-gift-given.html', {'gift1_given':gift1_given, 'gift2_given':gift2_given,
                'gift3_given':gift3_given,'gift4_given':gift4_given, 'gift_name_1':'Ba lô', 'gift_name_2':'Thùng 12 Lon', 'gift_name_3':'Nón', 'gift_name_4':'02 Lon Larue'} )
        
        if SP.brand == campain5:
            return render(request, 'report/alert-gift-given.html', {'gift1_given':gift1_given, 'gift2_given':gift2_given,
                'gift3_given':gift3_given,'gift4_given':gift4_given, 'gift_name_1':'Heineken Alu', 'gift_name_2':'Ba lô', 'gift_name_3':'Combo Thời Trang', 'gift_name_4':'Combo Du Lịch'} )
        
        if SP.brand == campain6:
            return render(request, 'report/alert-gift-given.html', {'gift1_given':gift1_given, 'gift2_given':gift2_given,
                'gift3_given':gift3_given,'gift4_given':gift4_given, 'gift_name_1':'Nón Strongbow', 'gift_name_2':'Túi Jute Bag', 'gift_name_3':'Túi Canvas ', 'gift_name_4':'Dù SB'})
        
        if SP.brand == campain1:
            return render(request, 'list_gift/alert-gift-given.html', {'gift1_given':gift1_given, 'gift2_given':gift2_given,
                'gift3_given':gift3_given,'gift4_given':gift4_given, 'gift5_given':gift5_given, 'gift6_given':gift6_given, 'gift_name_1':'Ly 30cl', 'gift_name_2':'Ly 33cl 3D', 'gift_name_3':'Ly Casablanca', 'gift_name_4':'Ví', 'gift_name_5':'Nón Tiger Crystal', 'gift_name_6':'Voucher Bia'}) 

        if SP.brand == campain2:
            return render(request, 'list_gift1/alert-gift-given.html', {'gift1_given':gift1_given, 'gift2_given':gift2_given,
                'gift3_given':gift3_given,'gift4_given':gift4_given, 'gift5_given':gift5_given, 'gift6_given':gift6_given, 'gift7_given':gift7_given,'gift_name_1':'Ly 30cl', 'gift_name_2':'Voucher beer', 'gift_name_3':'Festive Box', 'gift_name_4':'Túi du lịch Tiger', 'gift_name_5':'Loa Tiger', 'gift_name_6':'Ví Tiger ', 'gift_name_7':'Iphone 13'}) 

        if SP.brand == campain3:
            return render(request, 'list_gift2/alert-gift-given.html', {'gift1_given':gift1_given, 'gift2_given':gift2_given,
                'gift3_given':gift3_given,'gift4_given':gift4_given, 'gift5_given':gift5_given, 'gift6_given':gift6_given, 'gift7_given':gift7_given,'gift_name_1':'E-voucher 25k', 'gift_name_2':'E-voucher 50k', 'gift_name_3':'E-voucher 100k'})

        if SP.brand == campain4:
            return render(request, 'list_gift/alert-gift-given.html', {'gift1_given':gift1_given, 'gift2_given':gift2_given,
                'gift3_given':gift3_given,'gift4_given':gift4_given, 'gift5_given':gift5_given, 'gift6_given':gift6_given, 'gift_name_1':'Pin sạc', 'gift_name_2':'Ba lô', 'gift_name_3':'Bình Nước', 'gift_name_4':'Áo thun', 'gift_name_5':'Loa Bluetooth', 'gift_name_6':'Ly'}) 

        

#---------------------------------------GIFT REMAINING--------------------

login_required
def gift_remaining(request):
    try:
        SP = SalePerson.objects.get(user=request.user)
        report = giftReport.objects.get(created = datetime.date.today(), SP = request.user, outlet = SP.outlet)
        campain7 = Campain.objects.get(id=7)
        campain4 = Campain.objects.get(id=4)
        campain8 = Campain.objects.get(id=8)
        campain5 = Campain.objects.get(id=5)
        campain1 = Campain.objects.get(id=1)
        campain2 = Campain.objects.get(id=2)
        campain3 = Campain.objects.get(id=3)
        campain6 = Campain.objects.get(id=6)
        campain9 = Campain.objects.get(id=9)
        if SP.brand == campain7:
            return render(request,'report/listgift-remain.html', {'gift1_remaining':report.gift1_remaining,
                'gift2_remaining': report.gift2_remaining, 'gift3_remaining': report.gift3_remaining, 'gift4_remaining': report.gift4_remaining,
                'gift1_received':report.gift1_received, 'gift2_received':report.gift2_received,
                        'gift3_received':report.gift3_received, 'gift4_received':report.gift4_received,
                'gift1_given':report.gift1_given, 'gift2_given':report.gift2_given, 
                'gift3_given':report.gift3_given, 'gift4_given':report.gift4_given, 'gift_name_1':'Túi du lịch', 'gift_name_2':'Đồng Hồ Treo Tường', 'gift_name_3':'Bình Nước 1,6L', 'gift_name_4':'Ly'})
        if SP.brand == campain8:  
             return render(request,'report/listgift-remain.html', {'gift1_remaining':report.gift1_remaining,
                'gift2_remaining': report.gift2_remaining, 'gift3_remaining': report.gift3_remaining, 'gift4_remaining': report.gift4_remaining,
                'gift1_received':report.gift1_received, 'gift2_received':report.gift2_received,
                        'gift3_received':report.gift3_received, 'gift4_received':report.gift4_received,
                'gift1_given':report.gift1_given, 'gift2_given':report.gift2_given, 
                'gift3_given':report.gift3_given, 'gift4_given':report.gift4_given, 'gift_name_1':'Áo thun', 'gift_name_2':'Thùng 12 Lon', 'gift_name_3':'Nón', 'gift_name_4':'Ly'})
        
        if SP.brand == campain9:  
             return render(request,'report/listgift-remain.html', {'gift1_remaining':report.gift1_remaining,
                'gift2_remaining': report.gift2_remaining, 'gift3_remaining': report.gift3_remaining, 'gift4_remaining': report.gift4_remaining,
                'gift1_received':report.gift1_received, 'gift2_received':report.gift2_received,
                        'gift3_received':report.gift3_received, 'gift4_received':report.gift4_received,
                'gift1_given':report.gift1_given, 'gift2_given':report.gift2_given, 
                'gift3_given':report.gift3_given, 'gift4_given':report.gift4_given, 'gift_name_1':'Ba lô', 'gift_name_2':'Thùng 12 Lon', 'gift_name_3':'Nón', 'gift_name_4':'02 Lon Larue'})

        if SP.brand == campain5:  
             return render(request,'report/listgift-remain.html', {'gift1_remaining':report.gift1_remaining,
                'gift2_remaining': report.gift2_remaining, 'gift3_remaining': report.gift3_remaining, 'gift4_remaining': report.gift4_remaining,
                'gift1_received':report.gift1_received, 'gift2_received':report.gift2_received,
                        'gift3_received':report.gift3_received, 'gift4_received':report.gift4_received,
                'gift1_given':report.gift1_given, 'gift2_given':report.gift2_given, 
                'gift3_given':report.gift3_given, 'gift4_given':report.gift4_given,  'gift_name_1':'Heineken Alu', 'gift_name_2':'Ba lô', 'gift_name_3':'Combo Thời Trang', 'gift_name_4':'Combo Du Lịch'})
        
        if SP.brand == campain6:  
             return render(request,'report/listgift-remain.html', {'gift1_remaining':report.gift1_remaining,
                'gift2_remaining': report.gift2_remaining, 'gift3_remaining': report.gift3_remaining, 'gift4_remaining': report.gift4_remaining,
                'gift1_received':report.gift1_received, 'gift2_received':report.gift2_received,
                        'gift3_received':report.gift3_received, 'gift4_received':report.gift4_received,
                'gift1_given':report.gift1_given, 'gift2_given':report.gift2_given, 
                'gift3_given':report.gift3_given, 'gift4_given':report.gift4_given, 'gift_name_1':'Nón Strongbow', 'gift_name_2':'Túi Jute Bag', 'gift_name_3':'Túi Canvas ', 'gift_name_4':'Dù SB'})
        
        if SP.brand == campain1:
            return render(request,'list_gift/listgift-remain.html', {'gift1_remaining':report.gift1_remaining,
                'gift2_remaining': report.gift2_remaining, 'gift3_remaining': report.gift3_remaining, 'gift4_remaining': report.gift4_remaining,'gift5_remaining': report.gift5_remaining,'gift6_remaining': report.gift6_remaining,
                'gift1_received':report.gift1_received, 'gift2_received':report.gift2_received,
                        'gift3_received':report.gift3_received, 'gift4_received':report.gift4_received,'gift5_received':report.gift5_received, 'gift6_received':report.gift6_received,
                'gift1_given':report.gift1_given, 'gift2_given':report.gift2_given, 
                'gift3_given':report.gift3_given, 'gift4_given':report.gift4_given, 'gift5_given':report.gift5_given, 'gift6_given':report.gift6_given, 'gift_name_1':'Ly 30cl', 'gift_name_2':'Ly 33cl 3D', 'gift_name_3':'Ly Casablanca', 'gift_name_4':'Ví', 'gift_name_5':'Nón Tiger Crystal', 'gift_name_6':'Voucher Bia'})
        
        if SP.brand == campain2:
            return render(request,'list_gift1/listgift-remain.html', {'gift1_remaining':report.gift1_remaining,
                'gift2_remaining': report.gift2_remaining, 'gift3_remaining': report.gift3_remaining, 'gift4_remaining': report.gift4_remaining,'gift5_remaining': report.gift5_remaining,'gift6_remaining': report.gift6_remaining,'gift7_remaining': report.gift7_remaining,
                'gift1_received':report.gift1_received, 'gift2_received':report.gift2_received,
                        'gift3_received':report.gift3_received, 'gift4_received':report.gift4_received,'gift5_received':report.gift5_received, 'gift6_received':report.gift6_received, 'gift7_received':report.gift7_received,
                'gift1_given':report.gift1_given, 'gift2_given':report.gift2_given, 
                'gift3_given':report.gift3_given, 'gift4_given':report.gift4_given, 'gift5_given':report.gift5_given, 'gift6_given':report.gift6_given, 'gift7_given':report.gift7_given,'gift_name_1':'Ly 30cl', 'gift_name_2':'Voucher beer', 'gift_name_3':'Festive Box', 'gift_name_4':'Túi du lịch Tiger', 'gift_name_5':'Loa Tiger', 'gift_name_6':'Ví Tiger ', 'gift_name_7':'Iphone 13'})
        
        if SP.brand == campain3:
            return render(request,'list_gift2/listgift-remain.html', {'gift1_remaining':report.gift1_remaining,
                'gift2_remaining': report.gift2_remaining, 'gift3_remaining': report.gift3_remaining,
                'gift1_received':report.gift1_received, 'gift2_received':report.gift2_received,
                        'gift3_received':report.gift3_received,
                'gift1_given':report.gift1_given, 'gift2_given':report.gift2_given, 
                'gift3_given':report.gift3_given,'gift_name_1':'E-voucher 25k', 'gift_name_2':'E-voucher 50k', 'gift_name_3':'E-voucher 100k'})

        if SP.brand == campain4:
            return render(request,'list_gift/listgift-remain.html', {'gift1_remaining':report.gift1_remaining,
                'gift2_remaining': report.gift2_remaining, 'gift3_remaining': report.gift3_remaining, 'gift4_remaining': report.gift4_remaining,'gift5_remaining': report.gift5_remaining,'gift6_remaining': report.gift6_remaining,
                'gift1_received':report.gift1_received, 'gift2_received':report.gift2_received,
                        'gift3_received':report.gift3_received, 'gift4_received':report.gift4_received,'gift5_received':report.gift5_received, 'gift6_received':report.gift6_received,
                'gift1_given':report.gift1_given, 'gift2_given':report.gift2_given, 
                'gift3_given':report.gift3_given, 'gift4_given':report.gift4_given, 'gift5_given':report.gift5_given, 'gift6_given':report.gift6_given, 'gift_name_1':'Pin sạc', 'gift_name_2':'Ba lô', 'gift_name_3':'Bình Nước', 'gift_name_4':'Áo thun', 'gift_name_5':'Loa Bluetooth', 'gift_name_6':'Ly'})
    except:
        SP = SalePerson.objects.get(user=request.user)
        campain7 = Campain.objects.get(id=7)
        campain4 = Campain.objects.get(id=4)
        campain8 = Campain.objects.get(id=8)
        campain5 = Campain.objects.get(id=5)
        campain1 = Campain.objects.get(id=1)
        campain2 = Campain.objects.get(id=2)
        campain3 = Campain.objects.get(id=3)
        campain6 = Campain.objects.get(id=6)
        campain9 = Campain.objects.get(id=9)
        if SP.brand == campain7:
            return render(request,'report/listgift-remain.html', {'gift1_remaining':'0',
                'gift2_remaining': '0', 'gift3_remaining': '0', 'gift4_remaining': '0', 'gift_name_1':'Túi du lịch', 'gift_name_2':'Đồng Hồ Treo Tường', 'gift_name_3':'Bình Nước 1,6L', 'gift_name_4':'Ly'})
        if SP.brand == campain8:
            return render(request,'report/listgift-remain.html', {'gift1_remaining':'0',
                'gift2_remaining': '0', 'gift3_remaining': '0', 'gift4_remaining': '0', 'gift_name_1':'Áo thun', 'gift_name_2':'Thùng 12 Lon', 'gift_name_3':'Nón', 'gift_name_4':'Ly'})

        if SP.brand == campain9:
            return render(request,'report/listgift-remain.html', {'gift1_remaining':'0',
                'gift2_remaining': '0', 'gift3_remaining': '0', 'gift4_remaining': '0', 'gift_name_1':'Ba lô', 'gift_name_2':'Thùng 12 Lon', 'gift_name_3':'Nón', 'gift_name_4':'02 Lon Larue'})

        if SP.brand == campain5:
            return render(request,'report/listgift-remain.html', {'gift1_remaining':'0',
                'gift2_remaining': '0', 'gift3_remaining': '0', 'gift4_remaining': '0', 'gift_name_1':'Heineken Alu', 'gift_name_2':'Ba lô', 'gift_name_3':'Combo Thời Trang', 'gift_name_4':'Combo Du Lịch'})
        
        if SP.brand == campain6:
            return render(request,'report/listgift-remain.html', {'gift1_remaining':'0',
                'gift2_remaining': '0', 'gift3_remaining': '0', 'gift4_remaining': '0', 'gift_name_1':'Nón Strongbow', 'gift_name_2':'Túi Jute Bag', 'gift_name_3':'Túi Canvas ', 'gift_name_4':'Dù SB'})


        if SP.brand == campain1:
            return render(request,'list_gift/listgift-remain.html', {'gift1_remaining':'0',
                'gift2_remaining': '0', 'gift3_remaining': '0', 'gift4_remaining': '0', 'gift5_remaining': '0', 'gift6_remaining': '0', 'gift_name_1':'Ly 30cl', 'gift_name_2':'Ly 33cl 3D', 'gift_name_3':'Ly Casablanca', 'gift_name_4':'Ví', 'gift_name_5':'Nón Tiger Crystal', 'gift_name_6':'Voucher Bia'})
        
        if SP.brand == campain2:
            return render(request,'list_gift1/listgift-remain.html', {'gift1_remaining':'0',
                'gift2_remaining': '0', 'gift3_remaining': '0', 'gift4_remaining': '0', 'gift5_remaining': '0', 'gift6_remaining': '0', 'gift7_remaining': '0','gift_name_1':'Ly 30cl', 'gift_name_2':'Voucher beer', 'gift_name_3':'Festive Box', 'gift_name_4':'Túi du lịch Tiger', 'gift_name_5':'Loa Tiger', 'gift_name_6':'Ví Tiger ', 'gift_name_7':'Iphone 13'})
        
        if SP.brand == campain3:
            return render(request,'list_gift2/listgift-remain.html', {'gift1_remaining':'0',
                'gift2_remaining': '0', 'gift3_remaining': '0',
                'gift_name_1':'E-voucher 25k', 'gift_name_2':'E-voucher 50k', 'gift_name_3':'E-voucher 100k'})
  
        if SP.brand == campain4:
            return render(request,'list_gift/listgift-remain.html', {'gift1_remaining':'0',
                'gift2_remaining': '0', 'gift3_remaining': '0', 'gift4_remaining': '0', 'gift5_remaining': '0', 'gift6_remaining': '0', 'gift_name_1':'Pin sạc', 'gift_name_2':'Ba lô', 'gift_name_3':'Bình Nước', 'gift_name_4':'Áo thun', 'gift_name_5':'Loa Bluetooth', 'gift_name_6':'Ly'})




#----------------------------------------------------Report POSM-----------------------------------------------------

@csrf_exempt
@login_required
def reportPosm(request):
    user = request.user
    SP = SalePerson.objects.get(user=user)
    
    # image = request.POST.get('image')
    image = request.FILES.get('image')
    print(image.name)
    print(image.content_type)
    # print(image.read())
    report = posmReport.objects.filter(created = datetime.date.today(), SP = request.user, outlet = SP.outlet).count()
    if report < 1:
        posmReport.objects.create(image= image, SP=user, outlet=SP.outlet, campain=SP.brand)
        return JsonResponse({'created': 'true'})
    
    report = posmReport.objects.get(created = datetime.date.today(), SP = request.user, outlet = SP.outlet)
    report.image.delete()
    report.image=image
    report.save()
    print(report.image.url)
    return JsonResponse({'created': 'true'})


#--------------------------------------------------------------Report EndCase--------------------------------------------------

@login_required
def reportEndcase(request):
    SP = SalePerson.objects.get(user=request.user) 
    image = request.FILES.get('image')
    # report_posm = posmReport.objects.get(created = datetime.date.today(), SP = request.user, outlet = SP.outlet)
    # report_table = tableReport.objects.get(created = datetime.date.today(), SP = request.user, outlet = SP.outlet)
    # report_gift = giftReport.objects.get(created = datetime.date.today(), SP = request.user, outlet = SP.outlet)
    # report_consumer = consumerApproachReport.objects.get(created = datetime.date.today(), SP = request.user, outlet = SP.outlet)

    report = overallReport.objects.filter(created = datetime.date.today(), user = request.user, outlet = SP.outlet, campain=SP.brand).count()
    print(report)
    if report < 1:
        report = overallReport.objects.create(user = request.user, confirm = image, outlet = SP.outlet, campain=SP.brand)
    
        return JsonResponse({'created': 'true'})
    
    report = overallReport.objects.get(created = datetime.date.today(), user = request.user, outlet = SP.outlet)
    report.confirm.delete()
    report.confirm = image
    report.save()

    return JsonResponse({'created': 'true'})



def clean_data_today(request):
    report = overallReport.objects.filter(created = datetime.date.today()).delete()
    report1 = giftReport.objects.filter(created = datetime.date.today()).delete()
    report_table = tableReport.objects.filter(created = datetime.date.today()).delete()
    reportsale = report_sale.objects.filter(created = datetime.date.today()).delete()
    reportPOSM = posmReport.objects.filter(created = datetime.date.today()).delete()
    
    return HttpResponse('ok')


from datetime import datetime
def uploadFile_report(request, campainID):
	
            if "GET" == request.method:
                return render(request, 'report/upload-report.html', {'cam_id':campainID})
            else:
                excel_file = request.FILES["upload"]
                wb = openpyxl.load_workbook(excel_file)
                
                sheets = wb.sheetnames
                print(sheets[0])
                worksheet = wb[sheets[0]]   #Trang tính

                excel_data = list()
            
                for row in worksheet.iter_rows():
                    row_data = list()

                    for cell in row:
                        #if not (cell.value) == None: 
                        row_data.append(str(cell.value))

                    excel_data.append(row_data)
                print(len(excel_data)-2)
                # for i in range(len(excel_data)-1):
                #     a = Message(phone_number = '+84'+ excel_data[i+1][0][1:len(excel_data[i+1][0])], content = excel_data[i+1][1])
                #     a.save()
                list_outlet = []
                for i in range(len(excel_data)-1):
                    campain = Campain.objects.get(id=campainID)
                    #filter_outlet = outletInfo.objects.filter(compain=campain ,ouletID=excel_data[i+1][3], province=excel_data[i+1][2],  outlet_address=excel_data[i+1][6], outlet_Name=excel_data[i+1][7]).count()
                    filter_outlet = outletInfo.objects.filter(compain=campain ,ouletID=excel_data[i+1][2])
                    for outlet in filter_outlet:
                        campain = Campain.objects.get(id=campainID)
                        try:
                            date_filter = excel_data[i+1][0]
                            date_filter = datetime.strptime(date_filter,"%Y-%m-%d %H:%M:%S")
                        except:
                            date_filter = excel_data[i+1][0]
                            date_filter = datetime.strptime(date_filter,"%d/%m/%Y")
                        rp_table = tableReport.objects.filter(campain = campain,  outlet=outlet, created = date_filter)
                        rp_sale =  report_sale.objects.filter(campain=campain, outlet=outlet, created = date_filter)
                        consumers_rp = consumerApproachReport.objects.filter(campain=campain, outlet=outlet, created = date_filter)
                        list_gift_rp = giftReport.objects.filter(campain = campain,  outlet=outlet, created=date_filter)
                        for rp in rp_sale:
                            if excel_data[i+1][7] != None:
                                rp.beer_brand = str(excel_data[i+1][7])
                                rp.save()
                            if excel_data[i+1][8] != None:
                                rp.beer_HVN = str(excel_data[i+1][8])
                                rp.save()
                            if excel_data[i+1][9] != None:
                                rp.beer_other = str(excel_data[i+1][9])
                                rp.save()
                           
                        for rp in rp_table:
                            if excel_data[i+1][11] != None:
                                rp.brand_table = str(excel_data[i+1][11])
                                rp.save()
                            if excel_data[i+1][12] != None:
                                rp.HVN_table = str(excel_data[i+1][12])
                            
                            rp.other_beer_table = '0'
                            rp.other_table = '0'
                            rp.save()
                            #     rp.save()
                            # if excel_data[i+1][13] != None:
                            #     rp.other_beer_table = str(excel_data[i+1][13])
                            #     rp.save()
                            # if excel_data[i+1][14] != None:
                            #     rp.other_table = str(excel_data[i+1][14])
                            #     rp.save()
                        for rp in consumers_rp:
                            rp.Total_Consumers = '0'
                            rp.consumers_approach = '0'
                            rp.consumers_brough = '0'
                            # if excel_data[i+1][16] != None:
                            #     rp.Total_Consumers = str(excel_data[i+1][16])
                            #     rp.save()
                            # if excel_data[i+1][17] != None:
                            #     rp.consumers_approach = str(excel_data[i+1][17])
                            #     rp.save()
                            # if excel_data[i+1][19] != None:
                            #     rp.consumers_brough = str(excel_data[i+1][19])
                            #     rp.save()
                            rp.save()
                        for rp in list_gift_rp:
                            rp.gift1_received = '0'
                            rp.gift2_received = '0'
                            rp.gift3_received = '0'
                            rp.gift4_received= '0'
                            # if excel_data[i+1][21] != None or  excel_data[i+1][21] !='None' or  excel_data[i+1][21]!='':
                            #     rp.gift1_received = str(excel_data[i+1][21])
                            #     rp.save()
                            # if excel_data[i+1][22] != None:
                            #     rp.gift2_received = str(excel_data[i+1][22])
                            #     rp.save()
                            # if excel_data[i+1][23] != None:
                            #     rp.gift3_received = str(excel_data[i+1][23])
                            #     rp.save()
                            # if excel_data[i+1][24] != None:
                            #     rp.gift4_received = str(excel_data[i+1][24])
                            #     rp.save()
                            #rp.gift5_received = str(excel_data[i+1][25])
                            #rp.gift6_received = str(excel_data[i+1][26])
                            rp.gift1_given = '0'
                            rp.gift2_given = '0'
                            rp.gift3_given = '0'
                            rp.gift4_given = '0'
                            # if excel_data[i+1][27] != None or  excel_data[i+1][27] !='None' or  excel_data[i+1][27]!='':
                            #     rp.gift1_given = str(excel_data[i+1][27])
                            #     rp.save()
                            # if excel_data[i+1][28] != None:
                            #     rp.gift2_given = str(excel_data[i+1][28])
                            #     rp.save()
                            # if excel_data[i+1][29] != None:
                            #     rp.gift3_given = str(excel_data[i+1][29])
                            #     rp.save()
                            # if excel_data[i+1][30] != None:
                            #     rp.gift4_given = str(excel_data[i+1][30])
                            rp.save()
                            #rp.gift3_given = str(excel_data[i+1][31])
                            #rp.gift4_given = str(excel_data[i+1][32])
                            

                           
                            
                        print(date_filter)
                        # print(excel_data[i+1][1])
                        # print(excel_data[i+1][2])
                        # print(excel_data[i+1][3])
                        
                        # list_outlet.append(a)
            
                return render(request, "dashboard/management.html", {'list_outlet':list_outlet, "cam_id":campainID})
   
